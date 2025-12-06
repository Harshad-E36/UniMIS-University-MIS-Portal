from django.shortcuts import render, redirect
from .models import College, CollegeProgram, Taluka, District, Discipline, Programs, CollegeType, BelongsTo, academic_year, student_aggregate_master, staff_master_aggregate, UserCollege
from django.contrib.auth import get_user_model
from django.db.models import Prefetch
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from functools import wraps
from django.db import transaction
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import datetime
from io import BytesIO
import io


# Create your views here.

def ajax_login_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'message': 'Session expired. Please log in again.', 'status': 302})
        return view_func(request, *args, **kwargs)
    return _wrapped_view


def get_client_ip(request):
    """Get the real client IP address from request headers"""
    x_forwarded_for = request.META.get(('HTTP_X_FORWARDED_FOR'))
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    
    return ip

 # helper: determine user's college (None => admin / superuser)
def _get_user_college(user):
    if user.is_superuser:
        return None
    profile = getattr(user, "college_profile", None)
    if not profile:
        return None
    return profile.college


# Helper function to safely convert values to int
def _to_int(value, default=0):
    """Convert value to int safely, treating None/'' as default."""
    try:
        if value is None or value == "":
            return default
        return int(value)
    except (ValueError, TypeError):
        return default


def home(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    # Defaults (superuser / fallback)
    colleges_qs = College.objects.filter(is_deleted=False)
    disciplines_qs = Discipline.objects.all()
    programs_qs = Programs.objects.all()

    if not request.user.is_superuser:
        profile = UserCollege.objects.filter(user=request.user).first()
        if profile and profile.college and not profile.college.is_deleted:
            user_college = profile.college

            # only this college
            colleges_qs = College.objects.filter(id=user_college.id, is_deleted=False)
            # or: colleges_qs = College.objects.filter(pk=user_college.pk, is_deleted=False)

            # Get mapped disciplines + programs from CollegeProgram for this college
            cp_qs = CollegeProgram.objects.filter(College=user_college, is_deleted=False)

            # distinct discipline names and program names from mapping table
            discipline_name = cp_qs.values_list("Discipline", flat=True).distinct()
            program_name = cp_qs.values_list("ProgramName", flat=True).distinct()

            # Filter static tables using those names
            disciplines_qs = Discipline.objects.filter(DisciplineName__in=discipline_name)
           # IMPORTANT: filter programs by BOTH discipline and program name
            programs_qs = Programs.objects.filter(
                Discipline__DisciplineName__in=discipline_name,
                ProgramName__in=program_name,
            )


        else:
            # Normal user but no college assigned → send empty sets
            colleges_qs = College.objects.none()
            disciplines_qs = Discipline.objects.none()
            programs_qs = Programs.objects.none()

    return render(request, 'index.html', {
        "Colleges": colleges_qs,
        "disciplines": disciplines_qs,
        "Collegetype": CollegeType.objects.all(),
        "BelongsTo": BelongsTo.objects.all(),
        "programs": programs_qs,
        "academic_year": academic_year.objects.all(),
    })

def college_master(request):
    return render(request, 'college_master.html', {"disciplines" : Discipline.objects.all(), "Collegetype" : CollegeType.objects.all(), "BelongsTo": BelongsTo.objects.all()}) 


def student_master(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    CollegeCode = None
    if not request.user.is_superuser:
        profile = UserCollege.objects.filter(user=request.user).first()
        if profile:
            CollegeCode = profile.college.College_Code
            CollegeName = profile.college.College_Name
    return render(request, 'student_master.html', {"academic_year": academic_year.objects.all(), "college_code" : CollegeCode, "college_name": CollegeName})

def staff_master(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    CollegeCode = None
    if not request.user.is_superuser:
        profile = UserCollege.objects.filter(user=request.user).first()
        if profile:
            CollegeCode = profile.college.College_Code
            CollegeName = profile.college.College_Name
    return render(request, 'staff_master.html', {"academic_year": academic_year.objects.all(), "college_code" : CollegeCode, "college_name": CollegeName})


def user_status(request):
    if request.user.is_authenticated:
        response_data = {
            'is_authenticated' : True,
            'username' : request.user.username,
            'status' : 200,
        }
    else:
        response_data = {
            'is_authenticated' : False,
            'status' : 401
        }
    return JsonResponse(response_data)  


def signup(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        account_type = request.POST.get('account_type')

        if User.objects.filter(username = username).exists():
            response_data = {
                'message' : 'Username already exists',
                'status' : 400
            }
            return JsonResponse(response_data)
        
        if User.objects.filter(email = email).exists():
            response_data = {
                'message' : 'Email already exists',
                'status' : 400
            }
            return JsonResponse(response_data)
        
        user = User.objects.create_user(username=username, email=email, password=password)
        if account_type == "admin":
            user.is_staff = True
            user.is_superuser = True
        user.save()

        response_data = {
            'message' : 'User created successfully',
            'status' : 201
        }
        return JsonResponse(response_data)
    

def user_login(request):
    # If already logged in → go home
    if request.user.is_authenticated:
        return redirect('home')

    # If form is POST (login attempt)
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me',0)  # '1' if checked, 0 if not

        try:
            username = User.objects.get(email=email).username
        except User.DoesNotExist:
            return JsonResponse({'status': 401, 'message': 'Invalid email or password'})

        user = authenticate(request, username=username, password=password)


        if user:
            login(request, user)
            # Set session expiry based on "Remember Me"
            if remember_me == '1':
                request.session.set_expiry(21600)  # 2 weeks
            return JsonResponse({'status': 200, 'message': 'Login successful'})

        return JsonResponse({'status': 401, 'message': 'Invalid email or password'})

    # Show login page
    return render(request, 'login_page.html')
        
        
def user_logout(request):
    if request.method == 'POST':
        logout(request)
        response_data = {
            'message' : 'logout successful',
            'status' : 200
        }
        return JsonResponse(response_data)
    

@ajax_login_required
def add_edit_record(request):
    if request.method != "POST":
        return JsonResponse({"status": 400, "message": "Invalid method"}, status=400)

    # Safe ID parsing
    raw_id = request.POST.get('id', '').strip()
    try:
        record_id = int(raw_id) if raw_id else 0
    except ValueError:
        return JsonResponse({"status": 400, "message": "Invalid ID"}, status=400)

    college_code = request.POST.get('college_code')
    college_name = request.POST.get('college_name')
    address = request.POST.get('address')
    country = request.POST.get('country')
    state = request.POST.get('state')
    district = request.POST.get('district')
    taluka = request.POST.get('taluka')
    city = request.POST.get('city')
    pincode = request.POST.get('pincode')
    college_type = request.POST.get('college_type')
    belongs_to = request.POST.get('belongs_to')
    affiliated = request.POST.get('affiliated_to')

    assigned_user = request.POST.get('assigned_user_id')

    if not assigned_user:
        return JsonResponse({"error": "Please select a user to assign."}, status=400)
    
    if not college_code or not college_name:
        return JsonResponse({"error": "Missing college code or name."}, status=400)
    
    # get user
    try:
        user = User.objects.get(id=int(assigned_user))
    except (User.DoesNotExist, ValueError):
        return JsonResponse({"error": "Selected user does not exist."}, status=400)

    # Parse JSON safely
    programs_json = request.POST.get('disciplines_programs', '[]')
    try:
        programs = json.loads(programs_json)
    except Exception:
        return JsonResponse({"status": 400, "message": "Invalid programs JSON"}, status=400)

    # =============== EDIT MODE =============== 
    if record_id > 0:
        try:
            # change: make sure we don't edit a soft-deleted college
            college = College.objects.get(id=record_id, is_deleted=False)
        except College.DoesNotExist:
            return JsonResponse({"status": 404, "message": "Record not found"}, status=404)

        with transaction.atomic():

             # ---------- USER REASSIGNMENT LOGIC ----------
            # 1) Find any profiles currently pointing to this college
            existing_profiles = list(UserCollege.objects.filter(college=college))

            # 2) Get (or later create) profile for the selected user
            target_profile = UserCollege.objects.filter(user=user).first()

            # If this user is already assigned to a DIFFERENT college, block
            if target_profile and target_profile.college and target_profile.college != college:
                return JsonResponse(
                    {"error": "Selected user is already assigned to another college."},
                    status=409
                )

            # 3) Clear previous owners of this college (if any), except the selected user
            for p in existing_profiles:
                if p.user_id != user.id:
                    p.college = None
                    p.save(update_fields=['college'])

            # 4) Ensure selected user now owns this college
            if not target_profile:
                UserCollege.objects.create(user=user, college=college)
            else:
                target_profile.college = college
                target_profile.save(update_fields=['college'])


            # ---- Update college basic data ----
            college.College_Code = college_code
            college.College_Name = college_name
            college.address = address
            college.country = country
            college.state = state
            college.District = district
            college.taluka = taluka
            college.city = city
            college.pincode = pincode
            college.college_type = college_type
            college.belongs_to = belongs_to
            college.affiliated = affiliated
            college.updated_by = get_client_ip(request)
            college.save()

            # ---- Soft-delete all programs for this college first ----
            CollegeProgram.objects.filter(College=college).update(is_deleted=True)

            # Track active programs after this update
            active_program_ids = []

            # ---- Recreate / reactivate programs from request ----
            for item in programs:
                discipline = item.get('Discipline')
                program_name = item.get('ProgramName')

                if not discipline or not program_name:
                    # Skip invalid items
                    continue

                cp, created = CollegeProgram.objects.get_or_create(
                    College=college,
                    Discipline=discipline,
                    ProgramName=program_name,
                    defaults={'is_deleted': False}
                )

                # If it already exists but was soft-deleted, reactivate it
                if not created and cp.is_deleted:
                    cp.is_deleted = False
                    cp.save(update_fields=['is_deleted'])

                active_program_ids.append(cp.id)

            # ---- Sync student_aggregate_master with active programs ----
            if active_program_ids:
                # We only want to affect aggregates for removed programs,
                # and we must avoid unique constraint conflicts:
                # (College, Program, Academic_Year, is_deleted)
                to_soft_delete = student_aggregate_master.objects.filter(
                    College=college,
                    is_deleted=False
                ).exclude(
                    Program_id__in=active_program_ids
                )

                # Soft-delete one by one safely
                for agg in to_soft_delete.select_for_update():
                    conflict_qs = student_aggregate_master.objects.filter(
                        College=agg.College,
                        Program=agg.Program,
                        Academic_Year=agg.Academic_Year,
                        is_deleted=True
                    ).exclude(pk=agg.pk)

                    if conflict_qs.exists():
                        # There is an OLDER soft-deleted row.
                        old_soft_deleted = conflict_qs.first()

                        # HARD delete the old stale row (cleanup)
                        old_soft_deleted.delete()

                        # NOW safely soft-delete the active row
                        agg.is_deleted = True
                        agg.save(update_fields=['is_deleted'])
                    else:
                        # No duplicates → normal soft delete
                        agg.is_deleted = True
                        agg.save(update_fields=['is_deleted'])

            else:
                # No active programs left: soft-delete all aggregates for this college
                to_soft_delete = student_aggregate_master.objects.filter(
                    College=college,
                    is_deleted=False
                )

                for agg in to_soft_delete.select_for_update():
                    conflict_qs = student_aggregate_master.objects.filter(
                        College=agg.College,
                        Program=agg.Program,
                        Academic_Year=agg.Academic_Year,
                        is_deleted=True
                    ).exclude(pk=agg.pk)

                    if conflict_qs.exists():
                        agg.delete()
                    else:
                        agg.is_deleted = True
                        agg.save(update_fields=['is_deleted'])

        return JsonResponse({"status": 200, "message": "Record updated successfully"})

    # =============== ADD MODE =============== 

    # Check if this user is already assigned to some college
    UserExist = UserCollege.objects.filter(user=user).first()
    if UserExist and UserExist.college:
        return JsonResponse({"error": "User already assigned to a college."}, status=409)

    with transaction.atomic():
        college = College.objects.create(
            College_Code=college_code,
            College_Name=college_name,
            address=address,
            country=country,
            state=state,
            District=district,
            taluka=taluka,
            city=city,
            pincode=pincode,
            college_type=college_type,
            belongs_to=belongs_to,
            affiliated=affiliated,
            created_by=get_client_ip(request)
        )

        for item in programs:
            discipline = item.get('Discipline')
            program_name = item.get('ProgramName')

            if not discipline or not program_name:
                continue

            CollegeProgram.objects.create(
                College=college,
                Discipline=discipline,
                ProgramName=program_name
            )
        # (Usually no student_aggregate_master rows yet in ADD mode)
       
        # ---- Assign this college to the selected user ----
        if UserExist:
            # Profile exists but had no college assigned
            UserExist.college = college
            UserExist.save(update_fields=['college'])
        else:
            # No profile yet → create it
            UserCollege.objects.create(user=user, college=college)

    return JsonResponse({"status": 201, "message": "Record added successfully"})


@ajax_login_required
def delete_record(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        record = College.objects.get(id = id)
        record.is_deleted = True
        record.save()

        response_data = {
            'message' : 'record deleted successfully',
            'status' : 204
        }
        return JsonResponse(response_data)

@ajax_login_required
def get_dashboard_data(request):
    if request.method == 'GET':

         # helper: determine user's college (None => admin / superuser)
        
        def _get_user_college(user):
            if user.is_superuser:
                return None
            profile = UserCollege.objects.filter(user=user).first()
            if not profile:
                return None
            return profile.college

        user_college = _get_user_college(request.user)


        # If not superuser and no college assigned -> forbid
        if not request.user.is_superuser and not user_college:
            print("inside")
            response_data = {
                "detail": "No college assigned to this user. Contact admin.",
                 'status' :403
            }
            return JsonResponse(response_data)
                

        # get academic year from request
        academic_year = request.GET.get('year', None)

        # base queryset for students (filtered by academic_year if provided)
        student_qs = student_aggregate_master.objects.filter(is_deleted=False)
        staff_qs = staff_master_aggregate.objects.filter(is_deleted=False)
        if academic_year:
            student_qs = student_qs.filter(Academic_Year=academic_year)
            staff_qs = staff_qs.filter(Academic_Year=academic_year)

         # apply college restriction for normal user
        if user_college:
            student_qs = student_qs.filter(College=user_college)
            staff_qs = staff_qs.filter(College=user_college)

        # colleges aggregates
        if request.user.is_superuser:
            total_colleges = College.objects.filter(is_deleted=False).count()
        else:
            total_colleges = 1 if (user_college and not user_college.is_deleted) else 0

        total_students = student_qs.aggregate(total=Sum('total_students'))['total'] or 0
        total_staff = staff_qs.aggregate(total=Sum('total_staff'))['total'] or 0


        # helper function
        def stu_agg(field):
            return student_qs.aggregate(total= Sum(field))['total'] or 0
        
        def stf_agg(field):
            return staff_qs.aggregate(total = Sum(field))['total'] or 0
        
        def stu_washrooms_agg(field):
            return student_qs.aggregate(total=Sum(field))['total'] or 0
            
        def stf_washrooms_agg(field):
            return staff_qs.aggregate(total = Sum(field))['total'] or 0

        # student and staff aggregates
        response_data = {
            'total_colleges': total_colleges,
            'total_students': total_students,
            'total_staff' : total_staff,


            # students data
            'total_stu_washrooms': stu_washrooms_agg('total_washrooms'),
            'male_stu_washrooms': stu_washrooms_agg('male_washrooms'),
            'female_stu_washrooms': stu_washrooms_agg('female_washrooms'),

            'total_stu_male': stu_agg('total_male'),
            'total_stu_female': stu_agg('total_female'),
            'total_stu_others': stu_agg('total_others'),

            'open_stu_male': stu_agg('open_male'),
            'open_stu_female': stu_agg('open_female'),
            'open_stu_others': stu_agg('open_others'),

            'obc_stu_male': stu_agg('obc_male'),
            'obc_stu_female': stu_agg('obc_female'),
            'obc_stu_others': stu_agg('obc_others'),

            'sc_stu_male': stu_agg('sc_male'),
            'sc_stu_female': stu_agg('sc_female'),
            'sc_stu_others': stu_agg('sc_others'),

            'st_stu_male': stu_agg('st_male'),
            'st_stu_female': stu_agg('st_female'),
            'st_stu_others': stu_agg('st_others'),

            'nt_stu_male': stu_agg('nt_male'),
            'nt_stu_female': stu_agg('nt_female'),
            'nt_stu_others': stu_agg('nt_others'),

            'vjnt_stu_male': stu_agg('vjnt_male'),
            'vjnt_stu_female': stu_agg('vjnt_female'),
            'vjnt_stu_others': stu_agg('vjnt_others'),

            'ews_stu_male': stu_agg('ews_male'),
            'ews_stu_female': stu_agg('ews_female'),
            'ews_stu_others': stu_agg('ews_others'),

            'hindu_stu_male': stu_agg('hindu_male'),
            'hindu_stu_female': stu_agg('hindu_female'),
            'hindu_stu_others': stu_agg('hindu_others'),

            'muslim_stu_male': stu_agg('muslim_male'),
            'muslim_stu_female': stu_agg('muslim_female'),
            'muslim_stu_others': stu_agg('muslim_others'),

            'sikh_stu_male': stu_agg('sikh_male'),
            'sikh_stu_female': stu_agg('sikh_female'),
            'sikh_stu_others': stu_agg('sikh_others'),

            'christian_stu_male': stu_agg('christian_male'),
            'christian_stu_female': stu_agg('christian_female'),
            'christian_stu_others': stu_agg('christian_others'),

            'jain_stu_male': stu_agg('jain_male'),
            'jain_stu_female': stu_agg('jain_female'),
            'jain_stu_others': stu_agg('jain_others'),

            'buddhist_stu_male': stu_agg('buddhist_male'),
            'buddhist_stu_female': stu_agg('buddhist_female'),
            'buddhist_stu_others': stu_agg('buddhist_others'),

            'other_religion_stu_male': stu_agg('other_religion_male'),
            'other_religion_stu_female': stu_agg('other_religion_female'),
            'other_religion_stu_others': stu_agg('other_religion_others'),

            'no_disability_stu_male': stu_agg('no_disability_male'),
            'no_disability_stu_female': stu_agg('no_disability_female'),
            'no_disability_stu_others': stu_agg('no_disability_others'),

            'low_vision_stu_male': stu_agg('low_vision_male'),
            'low_vision_stu_female': stu_agg('low_vision_female'),
            'low_vision_stu_others': stu_agg('low_vision_others'),

            'blindness_stu_male': stu_agg('blindness_male'),
            'blindness_stu_female': stu_agg('blindness_female'),
            'blindness_stu_others': stu_agg('blindness_others'),

            'hearing_stu_male': stu_agg('hearing_male'),
            'hearing_stu_female': stu_agg('hearing_female'),
            'hearing_stu_others': stu_agg('hearing_others'),

            'locomotor_stu_male': stu_agg('locomotor_male'),
            'locomotor_stu_female': stu_agg('locomotor_female'),
            'locomotor_stu_others': stu_agg('locomotor_others'),

            'autism_stu_male': stu_agg('autism_male'),
            'autism_stu_female': stu_agg('autism_female'),
            'autism_stu_others': stu_agg('autism_others'),

            'other_disability_stu_male': stu_agg('other_disability_male'),
            'other_disability_stu_female': stu_agg('other_disability_female'),
            'other_disability_stu_others': stu_agg('other_disability_others'),

            # staff data
            'total_stf_washrooms': stf_washrooms_agg('total_washrooms'),
            'male_stf_washrooms': stf_washrooms_agg('male_washrooms'),
            'female_stf_washrooms': stf_washrooms_agg('female_washrooms'),

            'total_stf_male': stf_agg('total_male'),
            'total_stf_female': stf_agg('total_female'),
            'total_stf_others': stf_agg('total_others'),

            'open_stf_male': stf_agg('open_male'),
            'open_stf_female': stf_agg('open_female'),
            'open_stf_others': stf_agg('open_others'),

            'obc_stf_male': stf_agg('obc_male'),
            'obc_stf_female': stf_agg('obc_female'),
            'obc_stf_others': stf_agg('obc_others'),

            'sc_stf_male': stf_agg('sc_male'),
            'sc_stf_female': stf_agg('sc_female'),
            'sc_stf_others': stf_agg('sc_others'),

            'st_stf_male': stf_agg('st_male'),
            'st_stf_female': stf_agg('st_female'),
            'st_stf_others': stf_agg('st_others'),

            'nt_stf_male': stf_agg('nt_male'),
            'nt_stf_female': stf_agg('nt_female'),
            'nt_stf_others': stf_agg('nt_others'),

            'vjnt_stf_male': stf_agg('vjnt_male'),
            'vjnt_stf_female': stf_agg('vjnt_female'),
            'vjnt_stf_others': stf_agg('vjnt_others'),

            'ews_stf_male': stf_agg('ews_male'),
            'ews_stf_female': stf_agg('ews_female'),
            'ews_stf_others': stf_agg('ews_others'),

            'hindu_stf_male': stf_agg('hindu_male'),
            'hindu_stf_female': stf_agg('hindu_female'),
            'hindu_stf_others': stf_agg('hindu_others'),

            'muslim_stf_male': stf_agg('muslim_male'),
            'muslim_stf_female': stf_agg('muslim_female'),
            'muslim_stf_others': stf_agg('muslim_others'),

            'sikh_stf_male': stf_agg('sikh_male'),
            'sikh_stf_female': stf_agg('sikh_female'),
            'sikh_stf_others': stf_agg('sikh_others'),

            'christian_stf_male': stf_agg('christian_male'),
            'christian_stf_female': stf_agg('christian_female'),
            'christian_stf_others': stf_agg('christian_others'),

            'jain_stf_male': stf_agg('jain_male'),
            'jain_stf_female': stf_agg('jain_female'),
            'jain_stf_others': stf_agg('jain_others'),

            'buddhist_stf_male': stf_agg('buddhist_male'),
            'buddhist_stf_female': stf_agg('buddhist_female'),
            'buddhist_stf_others': stf_agg('buddhist_others'),

            'other_religion_stf_male': stf_agg('other_religion_male'),
            'other_religion_stf_female': stf_agg('other_religion_female'),
            'other_religion_stf_others': stf_agg('other_religion_others'),

            'no_disability_stf_male': stf_agg('no_disability_male'),
            'no_disability_stf_female': stf_agg('no_disability_female'),
            'no_disability_stf_others': stf_agg('no_disability_others'),

            'low_vision_stf_male': stf_agg('low_vision_male'),
            'low_vision_stf_female': stf_agg('low_vision_female'),
            'low_vision_stf_others': stf_agg('low_vision_others'),

            'blindness_stf_male': stf_agg('blindness_male'),
            'blindness_stf_female': stf_agg('blindness_female'),
            'blindness_stf_others': stf_agg('blindness_others'),

            'hearing_stf_male': stf_agg('hearing_male'),
            'hearing_stf_female': stf_agg('hearing_female'),
            'hearing_stf_others': stf_agg('hearing_others'),

            'locomotor_stf_male': stf_agg('locomotor_male'),
            'locomotor_stf_female': stf_agg('locomotor_female'),
            'locomotor_stf_others': stf_agg('locomotor_others'),

            'autism_stf_male': stf_agg('autism_male'),
            'autism_stf_female': stf_agg('autism_female'),
            'autism_stf_others': stf_agg('autism_others'),

            'other_disability_stf_male': stf_agg('other_disability_male'),
            'other_disability_stf_female': stf_agg('other_disability_female'),
            'other_disability_stf_others': stf_agg('other_disability_others'),

            'status': 200
        }
    return JsonResponse(response_data)


@ajax_login_required
def apply_filters(request):
    if request.method != "POST":
        return JsonResponse({"status": 400, "message": "Only POST allowed"})

    college_codes = request.POST.getlist('ColegeCode[]')
    college_names = request.POST.getlist('CollegeName[]')
    districts = request.POST.getlist('District[]')
    talukas = request.POST.getlist('Taluka[]')
    college_types = request.POST.getlist('CollegeType[]')
    belongs_tos = request.POST.getlist('BelongsTo[]')
    disciplines = request.POST.getlist('Discipline[]')
    programs = request.POST.getlist('Programs[]')
    academic_year = request.POST.get('year')  # matches your frontend

    def _zero_response():
          return JsonResponse({
            "status": 200,
            "message": "Filters applied successfully",
            "academic_year": academic_year,

            "total_colleges": 0,
            "total_students": 0,
            "total_staff" : 0,

            # washrooms
            "total_stu_washrooms": 0,
            "male_stu_washrooms": 0,
            "female_stu_washrooms": 0,

            # gender totals
            "total_stu_male": 0,
            "total_stu_female": 0,
            "total_stu_others": 0,

            # open category
            "open_stu_male": 0,
            "open_stu_female": 0,
            "open_stu_others": 0,

            # obc
            "obc_stu_male": 0,
            "obc_stu_female": 0,
            "obc_stu_others": 0,

            # sc
            "sc_stu_male": 0,
            "sc_stu_female": 0,
            "sc_stu_others": 0,

            # st
            "st_stu_male": 0,
            "st_stu_female": 0,
            "st_stu_others": 0,

            # nt
            "nt_stu_male": 0,
            "nt_stu_female": 0,
            "nt_stu_others": 0,

            # vjnt
            "vjnt_stu_male": 0,
            "vjnt_stu_female": 0,
            "vjnt_stu_others": 0,

            # ews
            "ews_stu_male": 0,
            "ews_stu_female": 0,
            "ews_stu_others": 0,

            # religions
            "hindu_stu_male": 0,
            "hindu_stu_female": 0,
            "hindu_stu_others": 0,

            "muslim_stu_male": 0,
            "muslim_stu_female": 0,
            "muslim_stu_others": 0,

            "sikh_stu_male": 0,
            "sikh_stu_female": 0,
            "sikh_stu_others": 0,

            "christian_stu_male": 0,
            "christian_stu_female": 0,
            "christian_stu_others": 0,

            "jain_stu_male": 0,
            "jain_stu_female": 0,
            "jain_stu_others": 0,

            "buddhist_stu_male": 0,
            "buddhist_stu_female": 0,
            "buddhist_stu_others": 0,

            "other_religion_stu_male": 0,
            "other_religion_stu_female": 0,
            "other_religion_stu_others": 0,

            # disabilities
            "no_disability_stu_male": 0,
            "no_disability_stu_female": 0,
            "no_disability_stu_others": 0,

            "low_vision_stu_male": 0,
            "low_vision_stu_female": 0,
            "low_vision_stu_others": 0,

            "blindness_stu_male": 0,
            "blindness_stu_female": 0,
            "blindness_stu_others": 0,

            "hearing_stu_male": 0,
            "hearing_stu_female": 0,
            "hearing_stu_others": 0,

            "locomotor_stu_male": 0,
            "locomotor_stu_female": 0,
            "locomotor_stu_others": 0,

            "autism_stu_male": 0,
            "autism_stu_female": 0,
            "autism_stu_others": 0,

            "other_disability_stu_male": 0,
            "other_disability_stu_female": 0,
            "other_disability_stu_others": 0,

            # staff data
            # staff data
            'total_stf_washrooms': 0,
            'male_stf_washrooms': 0,
            'female_stf_washrooms': 0,

            'total_stf_male': 0,
            'total_stf_female': 0,
            'total_stf_others': 0,

            'open_stf_male': 0,
            'open_stf_female': 0,
            'open_stf_others': 0,

            'obc_stf_male': 0,
            'obc_stf_female': 0,
            'obc_stf_others': 0,

            'sc_stf_male': 0,
            'sc_stf_female': 0,
            'sc_stf_others': 0,

            'st_stf_male': 0,
            'st_stf_female': 0,
            'st_stf_others': 0,

            'nt_stf_male': 0,
            'nt_stf_female': 0,
            'nt_stf_others': 0,

            'vjnt_stf_male': 0,
            'vjnt_stf_female': 0,
            'vjnt_stf_others': 0,

            'ews_stf_male': 0,
            'ews_stf_female': 0,
            'ews_stf_others': 0,

            'hindu_stf_male': 0,
            'hindu_stf_female': 0,
            'hindu_stf_others': 0,

            'muslim_stf_male': 0,
            'muslim_stf_female': 0,
            'muslim_stf_others': 0,

            'sikh_stf_male': 0,
            'sikh_stf_female': 0,
            'sikh_stf_others': 0,

            'christian_stf_male': 0,
            'christian_stf_female': 0,
            'christian_stf_others': 0,

            'jain_stf_male': 0,
            'jain_stf_female': 0,
            'jain_stf_others': 0,

            'buddhist_stf_male': 0,
            'buddhist_stf_female': 0,
            'buddhist_stf_others': 0,

            'other_religion_stf_male': 0,
            'other_religion_stf_female': 0,
            'other_religion_stf_others': 0,

            'no_disability_stf_male': 0,
            'no_disability_stf_female': 0,
            'no_disability_stf_others': 0,

            'low_vision_stf_male': 0,
            'low_vision_stf_female': 0,
            'low_vision_stf_others': 0,

            'blindness_stf_male': 0,
            'blindness_stf_female': 0,
            'blindness_stf_others': 0,

            'hearing_stf_male': 0,
            'hearing_stf_female': 0,
            'hearing_stf_others': 0,

            'locomotor_stf_male': 0,
            'locomotor_stf_female': 0,
            'locomotor_stf_others': 0,

            'autism_stf_male': 0,
            'autism_stf_female': 0,
            'autism_stf_others': 0,

            'other_disability_stf_male': 0,
            'other_disability_stf_female': 0,
            'other_disability_stf_others': 0,


            # special UI fields
            "colleges_without_student_data": [],
            "colleges_without_student_data_count": 0,
        })
    # ============================================================
    # BASE FILTER (on College master)
    # ============================================================
    filter_criteria = Q(is_deleted=False)
    if college_codes:
        filter_criteria &= Q(College_Code__in=college_codes)
    if college_names:
        filter_criteria &= Q(College_Name__in=college_names)
    if districts:
        filter_criteria &= Q(District__in=districts)
    if talukas:
        filter_criteria &= Q(taluka__in=talukas)
    if college_types:
        filter_criteria &= Q(college_type__in=college_types)
    if belongs_tos:
        filter_criteria &= Q(belongs_to__in=belongs_tos)

    base_ids = set(
        College.objects.filter(filter_criteria)
        .values_list("id", flat=True)
        .distinct()
    )

    # ============================================================
    # 🔒 Restrict by logged-in user's college (if not superuser)
    # ============================================================
    user = request.user
    if not user.is_superuser:
        profile = UserCollege.objects.filter(user=user).first()
        if not profile or not profile.college or profile.college.is_deleted:
            # Normal user but no valid college mapping → all zero response
            return _zero_response()

        # intersect filters with this user's single college
        user_college_id = profile.college.id
        base_ids = base_ids.intersection({user_college_id})

    
    # ============================================================
    # Determine which colleges match Discipline/Program using master table
    # (so colleges are counted even if no student rows exist for the selected year)
    # ============================================================
    if disciplines or programs:
        prog_master_q = Q(is_deleted=False)
        if disciplines:
            prog_master_q &= Q(Discipline__in=disciplines)
        if programs:
            prog_master_q &= Q(ProgramName__in=programs)

        prog_master_ids = set(
            CollegeProgram.objects.filter(prog_master_q)
            .values_list("College_id", flat=True)
            .distinct()
        )

        filtered_college_ids = list(base_ids.intersection(prog_master_ids))
    else:
        filtered_college_ids = list(base_ids)

    # ============================================================
    # If no matched colleges -> return zeros early (echo year)
    # ============================================================
    if not filtered_college_ids:
        return _zero_response()
      

    # ============================================================
    # Student rows to aggregate (year-specific, and narrowed by Program/Discipline if requested)
    # ============================================================
    students_qs = student_aggregate_master.objects.filter(
        College_id__in=filtered_college_ids,
        is_deleted=False
    )
    
    staff_qs = staff_master_aggregate.objects.filter(
        College_id__in=filtered_college_ids,
        is_deleted=False
    )
    if academic_year:
        students_qs = students_qs.filter(Academic_Year=academic_year)
        staff_qs = staff_qs.filter(Academic_Year=academic_year)

    # Narrow student rows by Program/Discipline for sums (this does NOT affect which colleges are counted)
    if disciplines:
        students_qs = students_qs.filter(Program__Discipline__in=disciplines)
        staff_qs = staff_qs.filter(Program__Discipline__in=disciplines)

    if programs:
        students_qs = students_qs.filter(Program__ProgramName__in=programs)
        staff_qs = staff_qs.filter(Program__ProgramName__in=programs)

    # Determine colleges that actually have student rows (for this filtered set & year)
    colleges_with_student_rows = set(
        student_aggregate_master.objects.filter(
            College_id__in=filtered_college_ids,
            is_deleted=False,
            **({"Academic_Year": academic_year} if academic_year else {})
        )
        .values_list("College_id", flat=True)
        .distinct()
    )

    # Determine colleges that actually have staff rows (for this filtered set & year)
    colleges_with_staff_rows = set(
        staff_master_aggregate.objects.filter(
            College_id__in=filtered_college_ids,
            is_deleted=False,
            **({"Academic_Year": academic_year} if academic_year else {})
        )
        .values_list("College_id", flat=True)
        .distinct()
    )

    colleges_without_student_data = list(set(filtered_college_ids) - colleges_with_student_rows)
    colleges_without_staff_data = list(set(filtered_college_ids) - colleges_with_staff_rows)


    # helper aggregator
    def stu_agg(field):
        return students_qs.aggregate(total= Sum(field))['total'] or 0
        
    def stf_agg(field):
        return staff_qs.aggregate(total = Sum(field))['total'] or 0
        
    def stu_washrooms_agg(field):
        return students_qs.aggregate(total=Sum(field))['total'] or 0
    
    def stf_washrooms_agg(field):
        return staff_qs.aggregate(total = Sum(field))['total'] or 0
    


         

    # ============================================================
    # Aggregations
    # ===========================================================
    # ============================================================
    # FINAL RESPONSE (includes colleges_without_student_data list)
    # ============================================================
    return JsonResponse({
        "status": 200,
        "message": "Filters applied successfully",
        "academic_year": academic_year,
        "total_colleges": len(filtered_college_ids),
        'total_students': stu_agg('total_students'),
        'total_staff' : stf_agg('total_staff'),


        # students data
        'total_stu_washrooms': stu_washrooms_agg('total_washrooms'),
        'male_stu_washrooms': stu_washrooms_agg('male_washrooms'),
        'female_stu_washrooms': stu_washrooms_agg('female_washrooms'),

        'total_stu_male': stu_agg('total_male'),
        'total_stu_female': stu_agg('total_female'),
        'total_stu_others': stu_agg('total_others'),

        'open_stu_male': stu_agg('open_male'),
        'open_stu_female': stu_agg('open_female'),
        'open_stu_others': stu_agg('open_others'),

        'obc_stu_male': stu_agg('obc_male'),
        'obc_stu_female': stu_agg('obc_female'),
        'obc_stu_others': stu_agg('obc_others'),

        'sc_stu_male': stu_agg('sc_male'),
        'sc_stu_female': stu_agg('sc_female'),
        'sc_stu_others': stu_agg('sc_others'),

        'st_stu_male': stu_agg('st_male'),
        'st_stu_female': stu_agg('st_female'),
        'st_stu_others': stu_agg('st_others'),

        'nt_stu_male': stu_agg('nt_male'),
        'nt_stu_female': stu_agg('nt_female'),
        'nt_stu_others': stu_agg('nt_others'),

        'vjnt_stu_male': stu_agg('vjnt_male'),
        'vjnt_stu_female': stu_agg('vjnt_female'),
        'vjnt_stu_others': stu_agg('vjnt_others'),

        'ews_stu_male': stu_agg('ews_male'),
        'ews_stu_female': stu_agg('ews_female'),
        'ews_stu_others': stu_agg('ews_others'),

        'hindu_stu_male': stu_agg('hindu_male'),
        'hindu_stu_female': stu_agg('hindu_female'),
        'hindu_stu_others': stu_agg('hindu_others'),

        'muslim_stu_male': stu_agg('muslim_male'),
        'muslim_stu_female': stu_agg('muslim_female'),
        'muslim_stu_others': stu_agg('muslim_others'),

        'sikh_stu_male': stu_agg('sikh_male'),
        'sikh_stu_female': stu_agg('sikh_female'),
        'sikh_stu_others': stu_agg('sikh_others'),

        'christian_stu_male': stu_agg('christian_male'),
        'christian_stu_female': stu_agg('christian_female'),
        'christian_stu_others': stu_agg('christian_others'),

        'jain_stu_male': stu_agg('jain_male'),
        'jain_stu_female': stu_agg('jain_female'),
        'jain_stu_others': stu_agg('jain_others'),

        'buddhist_stu_male': stu_agg('buddhist_male'),
        'buddhist_stu_female': stu_agg('buddhist_female'),
        'buddhist_stu_others': stu_agg('buddhist_others'),

        'other_religion_stu_male': stu_agg('other_religion_male'),
        'other_religion_stu_female': stu_agg('other_religion_female'),
        'other_religion_stu_others': stu_agg('other_religion_others'),

        'no_disability_stu_male': stu_agg('no_disability_male'),
        'no_disability_stu_female': stu_agg('no_disability_female'),
        'no_disability_stu_others': stu_agg('no_disability_others'),

        'low_vision_stu_male': stu_agg('low_vision_male'),
        'low_vision_stu_female': stu_agg('low_vision_female'),
        'low_vision_stu_others': stu_agg('low_vision_others'),

        'blindness_stu_male': stu_agg('blindness_male'),
        'blindness_stu_female': stu_agg('blindness_female'),
        'blindness_stu_others': stu_agg('blindness_others'),

        'hearing_stu_male': stu_agg('hearing_male'),
        'hearing_stu_female': stu_agg('hearing_female'),
        'hearing_stu_others': stu_agg('hearing_others'),

        'locomotor_stu_male': stu_agg('locomotor_male'),
        'locomotor_stu_female': stu_agg('locomotor_female'),
        'locomotor_stu_others': stu_agg('locomotor_others'),

        'autism_stu_male': stu_agg('autism_male'),
        'autism_stu_female': stu_agg('autism_female'),
        'autism_stu_others': stu_agg('autism_others'),

        'other_disability_stu_male': stu_agg('other_disability_male'),
        'other_disability_stu_female': stu_agg('other_disability_female'),
        'other_disability_stu_others': stu_agg('other_disability_others'),

        # staff data
        'total_stf_washrooms': stf_washrooms_agg('total_washrooms'),
        'male_stf_washrooms': stf_washrooms_agg('male_washrooms'),
        'female_stf_washrooms': stf_washrooms_agg('female_washrooms'),

        'total_stf_male': stf_agg('total_male'),
        'total_stf_female': stf_agg('total_female'),
        'total_stf_others': stf_agg('total_others'),

        'open_stf_male': stf_agg('open_male'),
        'open_stf_female': stf_agg('open_female'),
        'open_stf_others': stf_agg('open_others'),

        'obc_stf_male': stf_agg('obc_male'),
        'obc_stf_female': stf_agg('obc_female'),
        'obc_stf_others': stf_agg('obc_others'),

        'sc_stf_male': stf_agg('sc_male'),
        'sc_stf_female': stf_agg('sc_female'),
        'sc_stf_others': stf_agg('sc_others'),

        'st_stf_male': stf_agg('st_male'),
        'st_stf_female': stf_agg('st_female'),
        'st_stf_others': stf_agg('st_others'),

        'nt_stf_male': stf_agg('nt_male'),
        'nt_stf_female': stf_agg('nt_female'),
        'nt_stf_others': stf_agg('nt_others'),

        'vjnt_stf_male': stf_agg('vjnt_male'),
        'vjnt_stf_female': stf_agg('vjnt_female'),
        'vjnt_stf_others': stf_agg('vjnt_others'),

        'ews_stf_male': stf_agg('ews_male'),
        'ews_stf_female': stf_agg('ews_female'),
        'ews_stf_others': stf_agg('ews_others'),

        'hindu_stf_male': stf_agg('hindu_male'),
        'hindu_stf_female': stf_agg('hindu_female'),
        'hindu_stf_others': stf_agg('hindu_others'),

        'muslim_stf_male': stf_agg('muslim_male'),
        'muslim_stf_female': stf_agg('muslim_female'),
        'muslim_stf_others': stf_agg('muslim_others'),

        'sikh_stf_male': stf_agg('sikh_male'),
        'sikh_stf_female': stf_agg('sikh_female'),
        'sikh_stf_others': stf_agg('sikh_others'),

        'christian_stf_male': stf_agg('christian_male'),
        'christian_stf_female': stf_agg('christian_female'),
        'christian_stf_others': stf_agg('christian_others'),

        'jain_stf_male': stf_agg('jain_male'),
        'jain_stf_female': stf_agg('jain_female'),
        'jain_stf_others': stf_agg('jain_others'),

        'buddhist_stf_male': stf_agg('buddhist_male'),
        'buddhist_stf_female': stf_agg('buddhist_female'),
        'buddhist_stf_others': stf_agg('buddhist_others'),

        'other_religion_stf_male': stf_agg('other_religion_male'),
        'other_religion_stf_female': stf_agg('other_religion_female'),
        'other_religion_stf_others': stf_agg('other_religion_others'),

        'no_disability_stf_male': stf_agg('no_disability_male'),
        'no_disability_stf_female': stf_agg('no_disability_female'),
        'no_disability_stf_others': stf_agg('no_disability_others'),

        'low_vision_stf_male': stf_agg('low_vision_male'),
        'low_vision_stf_female': stf_agg('low_vision_female'),
        'low_vision_stf_others': stf_agg('low_vision_others'),

        'blindness_stf_male': stf_agg('blindness_male'),
        'blindness_stf_female': stf_agg('blindness_female'),
        'blindness_stf_others': stf_agg('blindness_others'),

        'hearing_stf_male': stf_agg('hearing_male'),
        'hearing_stf_female': stf_agg('hearing_female'),
        'hearing_stf_others': stf_agg('hearing_others'),

        'locomotor_stf_male': stf_agg('locomotor_male'),
        'locomotor_stf_female': stf_agg('locomotor_female'),
        'locomotor_stf_others': stf_agg('locomotor_others'),

        'autism_stf_male': stf_agg('autism_male'),
        'autism_stf_female': stf_agg('autism_female'),
        'autism_stf_others': stf_agg('autism_others'),

        'other_disability_stf_male': stf_agg('other_disability_male'),
        'other_disability_stf_female': stf_agg('other_disability_female'),
        'other_disability_stf_others': stf_agg('other_disability_others'),

        # new fields to help UI flag missing student data per-college
        "colleges_without_student_data": colleges_without_student_data,
        "colleges_without_student_data_count": len(colleges_without_student_data)
    })


def get_talukas(request):
    if request.method == "GET":
        district_name = request.GET.get('district')
        if not district_name:
            return JsonResponse({'talukas': []})
        
        talukas = Taluka.objects.filter(District__DistrictName=district_name).values_list('TalukaName', flat=True)
        return JsonResponse({'talukas': list(talukas)})
    return JsonResponse({'talukas': []})


def get_programs_for_discipline(request):
    disciplines = request.GET.getlist('discipline')
    response_data = {}

    for discipline_name in disciplines:
        # Fetch all programs related to this discipline from DB
        programs = Programs.objects.filter(
            Discipline__DisciplineName=discipline_name
        ).values_list('ProgramName', flat=True)

        # If no programs found, use default placeholder
        if programs:
            response_data[discipline_name] = list(programs)
        else:
            response_data[discipline_name] = ["No programs available"]

    return JsonResponse(response_data)


@ajax_login_required
def get_college_data_for_student_and_staff_modal(request):
    if request.method != "GET":
        return JsonResponse({'status': 400, 'message': 'Invalid request'})

    college_code = request.GET.get('college_code')
    academic_year = request.GET.get('academic_year')
    mode = request.GET.get('mode', 'add')
    page = request.GET.get('page')

    if not college_code:
        return JsonResponse({'status': 400, 'message': 'Missing college_code'})

    try:
        college = College.objects.get(College_Code=college_code, is_deleted=False)
    except College.DoesNotExist:
        return JsonResponse({'status': 404, 'message': 'College not found'})

    # ---------- Build discipline -> sorted(programs) map ----------
    all_programs = (
        CollegeProgram.objects
        .filter(College=college, is_deleted=False)
        .values("Discipline", "ProgramName")
        .order_by("Discipline", "ProgramName")             # <-- alphabetical sorting HERE
    )

    discipline_map = {}
    for item in all_programs:
        disc = item.get("Discipline") or "Unspecified"
        prog = item.get("ProgramName") or "Unnamed Program"
        discipline_map.setdefault(disc, []).append(prog)

    # Ensure alphabetical discipline ordering
    discipline_map = {
        disc: sorted(progs, key=lambda x: x.lower())
        for disc, progs in sorted(discipline_map.items(), key=lambda x: x[0].lower())
    }

    full_address = f"{college.address}, {college.taluka}, {college.District} - {college.pincode}"

    base_college_data = {
        "College_Code": college.College_Code,
        "College_Name": college.College_Name,
        "address": full_address,
        "District": college.District,
        "Taluka": college.taluka,
        "pincode": college.pincode,
        "country": college.country,
        "state": college.state,
        "affiliated": college.affiliated,
        "programs": discipline_map
    }

    # ADD MODE
    if mode == "add" and (page =="student" or page == "staff"):
        return JsonResponse({
            "status": 200,
            "mode": "add",
            "academic_year": academic_year,
            "college_data": base_college_data,
            "records": {}
        })

    # EDIT MODE
    if mode == "edit":
        if page == "student":

            if not academic_year:
                return JsonResponse({'status': 400, 'message': 'Missing academic_year'})

            aggregates = student_aggregate_master.objects.filter(
                College=college,
                Academic_Year=academic_year,
                is_deleted=False
            ).select_related("Program")

            if not aggregates.exists():
                return JsonResponse({
                    "status": 404,
                    "message": "No records found for this year"
                })

            filled_records = {}

            for agg in aggregates:
                prog_name = agg.Program.ProgramName if agg.Program else f"program_{agg.pk}"

                filled_records[prog_name] = {
                    "total_students": agg.total_students,
                    "gender": {
                        "male": agg.total_male,
                        "female": agg.total_female,
                        "others": agg.total_others,
                    },
                    "washrooms": {
                        "total_washrooms":agg.total_washrooms or 0,
                        "male_washrooms": agg.male_washrooms or 0,
                        "female_washrooms" : agg.female_washrooms or 0,
                    },
                    "category": {
                        "open": {
                            "male": agg.open_male,
                            "female": agg.open_female,
                            "others": agg.open_others,
                        },
                        "obc": {
                            "male": agg.obc_male,
                            "female": agg.obc_female,
                            "others": agg.obc_others,
                        },
                        "sc": {
                            "male": agg.sc_male,
                            "female": agg.sc_female,
                            "others": agg.sc_others,
                        },
                        "st": {
                            "male": agg.st_male,
                            "female": agg.st_female,
                            "others": agg.st_others,
                        },
                        "nt": {
                            "male": agg.nt_male,
                            "female": agg.nt_female,
                            "others": agg.nt_others,
                        },
                        "vjnt": {
                            "male": agg.vjnt_male,
                            "female": agg.vjnt_female,
                            "others": agg.vjnt_others,
                        },
                        "ews": {
                            "male": agg.ews_male,
                            "female": agg.ews_female,
                            "others": agg.ews_others,
                        },
                    },
                    "religion": {
                        "hindu": {
                            "male": agg.hindu_male,
                            "female": agg.hindu_female,
                            "others": agg.hindu_others,
                        },
                        "muslim": {
                            "male": agg.muslim_male,
                            "female": agg.muslim_female,
                            "others": agg.muslim_others,
                        },
                        "sikh": {
                            "male": agg.sikh_male,
                            "female": agg.sikh_female,
                            "others": agg.sikh_others,
                        },
                        "christian": {
                            "male": agg.christian_male,
                            "female": agg.christian_female,
                            "others": agg.christian_others,
                        },
                        "jain": {
                            "male": agg.jain_male,
                            "female": agg.jain_female,
                            "others": agg.jain_others,
                        },
                        "buddhist": {
                            "male": agg.buddhist_male,
                            "female": agg.buddhist_female,
                            "others": agg.buddhist_others,
                        },
                        "other_religion": {
                            "male": agg.other_religion_male,
                            "female": agg.other_religion_female,
                            "others": agg.other_religion_others,
                        },
                    },
                    "disability": {
                        "no_disability": {
                            "male": agg.no_disability_male,
                            "female": agg.no_disability_female,
                            "others": agg.no_disability_others,
                        },
                        "lowvision": {
                            "male": agg.low_vision_male,
                            "female": agg.low_vision_female,
                            "others": agg.low_vision_others,
                        },
                        "blindness": {
                            "male": agg.blindness_male,
                            "female": agg.blindness_female,
                            "others": agg.blindness_others,
                        },
                        "hearing": {
                            "male": agg.hearing_male,
                            "female": agg.hearing_female,
                            "others": agg.hearing_others,
                        },
                        "locomotor": {
                            "male": agg.locomotor_male,
                            "female": agg.locomotor_female,
                            "others": agg.locomotor_others,
                        },
                        "autism": {
                            "male": agg.autism_male,
                            "female": agg.autism_female,
                            "others": agg.autism_others,
                        },
                        "other_disability": {
                            "male": agg.other_disability_male,
                            "female": agg.other_disability_female,
                            "others": agg.other_disability_others,
                        },
                    }
                }

            return JsonResponse({
                "status": 200,
                "mode": "edit",
                "academic_year": academic_year,
                "college_data": base_college_data,
                "records": filled_records
            })
        
        elif page == "staff":
            if not academic_year:
                return JsonResponse({'status': 400, 'message': 'Missing academic_year'})

            aggregates = staff_master_aggregate.objects.filter(
                College=college,
                Academic_Year=academic_year,
                is_deleted=False
            ).select_related("Program")

            if not aggregates.exists():
                return JsonResponse({
                    "status": 404,
                    "message": "No records found for this year"
                })

            filled_records = {}

            for agg in aggregates:
                prog_name = agg.Program.ProgramName if agg.Program else f"program_{agg.pk}"

                filled_records[prog_name] = {
                    "total_students": agg.total_staff,
                    "gender": {
                        "male": agg.total_male,
                        "female": agg.total_female,
                        "others": agg.total_others,
                    },
                    "washrooms": {
                        "total_washrooms":agg.total_washrooms or 0,
                        "male_washrooms": agg.male_washrooms or 0,
                        "female_washrooms" : agg.female_washrooms or 0,
                    },
                    "category": {
                        "open": {
                            "male": agg.open_male,
                            "female": agg.open_female,
                            "others": agg.open_others,
                        },
                        "obc": {
                            "male": agg.obc_male,
                            "female": agg.obc_female,
                            "others": agg.obc_others,
                        },
                        "sc": {
                            "male": agg.sc_male,
                            "female": agg.sc_female,
                            "others": agg.sc_others,
                        },
                        "st": {
                            "male": agg.st_male,
                            "female": agg.st_female,
                            "others": agg.st_others,
                        },
                        "nt": {
                            "male": agg.nt_male,
                            "female": agg.nt_female,
                            "others": agg.nt_others,
                        },
                        "vjnt": {
                            "male": agg.vjnt_male,
                            "female": agg.vjnt_female,
                            "others": agg.vjnt_others,
                        },
                        "ews": {
                            "male": agg.ews_male,
                            "female": agg.ews_female,
                            "others": agg.ews_others,
                        },
                    },
                    "religion": {
                        "hindu": {
                            "male": agg.hindu_male,
                            "female": agg.hindu_female,
                            "others": agg.hindu_others,
                        },
                        "muslim": {
                            "male": agg.muslim_male,
                            "female": agg.muslim_female,
                            "others": agg.muslim_others,
                        },
                        "sikh": {
                            "male": agg.sikh_male,
                            "female": agg.sikh_female,
                            "others": agg.sikh_others,
                        },
                        "christian": {
                            "male": agg.christian_male,
                            "female": agg.christian_female,
                            "others": agg.christian_others,
                        },
                        "jain": {
                            "male": agg.jain_male,
                            "female": agg.jain_female,
                            "others": agg.jain_others,
                        },
                        "buddhist": {
                            "male": agg.buddhist_male,
                            "female": agg.buddhist_female,
                            "others": agg.buddhist_others,
                        },
                        "other_religion": {
                            "male": agg.other_religion_male,
                            "female": agg.other_religion_female,
                            "others": agg.other_religion_others,
                        },
                    },
                    "disability": {
                        "no_disability": {
                            "male": agg.no_disability_male,
                            "female": agg.no_disability_female,
                            "others": agg.no_disability_others,
                        },
                        "lowvision": {
                            "male": agg.low_vision_male,
                            "female": agg.low_vision_female,
                            "others": agg.low_vision_others,
                        },
                        "blindness": {
                            "male": agg.blindness_male,
                            "female": agg.blindness_female,
                            "others": agg.blindness_others,
                        },
                        "hearing": {
                            "male": agg.hearing_male,
                            "female": agg.hearing_female,
                            "others": agg.hearing_others,
                        },
                        "locomotor": {
                            "male": agg.locomotor_male,
                            "female": agg.locomotor_female,
                            "others": agg.locomotor_others,
                        },
                        "autism": {
                            "male": agg.autism_male,
                            "female": agg.autism_female,
                            "others": agg.autism_others,
                        },
                        "other_disability": {
                            "male": agg.other_disability_male,
                            "female": agg.other_disability_female,
                            "others": agg.other_disability_others,
                        },
                    }
                }

            return JsonResponse({
                "status": 200,
                "mode": "edit",
                "academic_year": academic_year,
                "college_data": base_college_data,
                "records": filled_records
            })

    return JsonResponse({"status": 400, "message": "Invalid mode"})


@ajax_login_required
def add_student_aggregate(request):
    if request.method == "POST":
        try:
            payload = json.loads(request.body)
        except json.JSONDecodeError:
            return HttpResponseBadRequest("Invalid JSON")
        
        college_code = payload.get('college_code')
        academic_year = payload.get('academic_year')
        records = payload.get('records', [])

        if not college_code or not academic_year:
            return JsonResponse({'status': 400, 'message': 'Missing required fields'})

        try:
            college = College.objects.get(College_Code=college_code, is_deleted=False)      
        except College.DoesNotExist:
            return JsonResponse({'status': 404, 'message': 'College not found'})
        
        if not isinstance(records, dict) or len(records) == 0:
            return JsonResponse({"status": 400, "message": "No records provided"}, status=400)

        saved = []
        errors = []

        
        with transaction.atomic():
            for program_name, data in records.items():
                program_obj = CollegeProgram.objects.filter(
                    College=college,
                    ProgramName=program_name,
                    is_deleted=False
                ).first()
                
                if not program_obj:
                    errors.append({
                        "program": program_name,
                        "error": f"Program '{program_name}' not found for college '{college_code}'"
                    })
                    continue
                
                # ---- parse numeric fields safely ----
                total_students = _to_int(data.get("total_students"), 0)

                gender = data.get("gender", {}) or {}
                male = _to_int(gender.get("male"), 0)
                female = _to_int(gender.get("female"), 0)
                others = _to_int(gender.get("others") or gender.get("other"), 0)

                washrooms = data.get("washrooms", {}) or {}
                total_washrooms = _to_int(washrooms.get("total"), 0)
                male_washrooms = _to_int(washrooms.get("male"), 0)
                female_washrooms = _to_int(washrooms.get("female"), 0)


                category = data.get("category", {}) or {}
                open = category.get("open", {}) or {}
                open_male = _to_int(open.get("male"), 0)
                open_female = _to_int(open.get("female"), 0)
                open_others = _to_int(open.get("others") or open.get("other"), 0)

                obc = category.get("obc", {}) or {}
                obc_male = _to_int(obc.get("male"), 0)
                obc_female = _to_int(obc.get("female"), 0)
                obc_others = _to_int(obc.get("others") or obc.get("other"), 0)
                
                sc = category.get("sc", {}) or {}
                sc_male = _to_int(sc.get("male"), 0)
                sc_female = _to_int(sc.get("female"), 0)
                sc_others = _to_int(sc.get("others") or sc.get("other"), 0)

                st = category.get("st", {}) or {}
                st_male = _to_int(st.get("male"), 0)
                st_female = _to_int(st.get("female"), 0)
                st_others = _to_int(st.get("others") or st.get("other"), 0) 

                nt = category.get("nt", {}) or {}
                nt_male = _to_int(nt.get("male"), 0)
                nt_female = _to_int(nt.get("female"), 0)
                nt_others = _to_int(nt.get("others") or nt.get("other"), 0)

                vjnt = category.get("vjnt", {}) or {}
                vjnt_male = _to_int(vjnt.get("male"), 0)
                vjnt_female = _to_int(vjnt.get("female"), 0)
                vjnt_others = _to_int(vjnt.get("others") or vjnt.get("other"), 0)

                ews = category.get("ews", {}) or {}
                ews_male = _to_int(ews.get("male"), 0)
                ews_female = _to_int(ews.get("female"), 0)
                ews_others = _to_int(ews.get("others") or ews.get("other"), 0)


                religion = data.get("religion", {}) or {}

                hindu = religion.get("hindu", {}) or {}
                hindu_male = _to_int(hindu.get("male"), 0)
                hindu_female = _to_int(hindu.get("female"), 0)
                hindu_others = _to_int(hindu.get("others") or hindu.get("other"), 0)

                muslim = religion.get("muslim", {}) or {}
                muslim_male = _to_int(muslim.get("male"), 0)
                muslim_female = _to_int(muslim.get("female"), 0)
                muslim_others = _to_int(muslim.get("others") or muslim.get("other"), 0)

                sikh = religion.get("sikh", {}) or {}
                sikh_male = _to_int(sikh.get("male"), 0)
                sikh_female = _to_int(sikh.get("female"), 0)
                sikh_others = _to_int(sikh.get("others") or sikh.get("other"), 0)

                christian = religion.get("christian", {}) or {}
                christian_male = _to_int(christian.get("male"), 0)
                christian_female = _to_int(christian.get("female"), 0)
                christian_others = _to_int(christian.get("others") or christian.get("other"), 0)

                jain = religion.get("jain", {}) or {}
                jain_male = _to_int(jain.get("male"), 0)
                jain_female = _to_int(jain.get("female"), 0)
                jain_others = _to_int(jain.get("others") or jain.get("other"), 0)

                buddhist = religion.get("buddhist", {}) or {}
                buddhist_male = _to_int(buddhist.get("male"), 0)
                buddhist_female = _to_int(buddhist.get("female"), 0)
                buddhist_others = _to_int(buddhist.get("others") or buddhist.get("other"), 0)

                other_religion = religion.get("other_religion", {}) or {}
                other_religion_male = _to_int(other_religion.get("male"), 0)
                other_religion_female = _to_int(other_religion.get("female"), 0)
                other_religion_others = _to_int(other_religion.get("others") or other_religion.get("other"), 0)


                dis = data.get("disability", {}) or {}
                no_disability = dis.get("no_disability", {}) or {}
                no_disability_male = _to_int(no_disability.get("male"), 0)
                no_disability_female = _to_int(no_disability.get("female"), 0)
                no_disability_others = _to_int(no_disability.get("others") or no_disability.get("other"), 0)

                low_vision = dis.get("lowvision", {}) or {}
                low_vision_male = _to_int(low_vision.get("male"), 0)
                low_vision_female = _to_int(low_vision.get("female"), 0)
                low_vision_others = _to_int(low_vision.get("others") or low_vision.get("other"), 0)

                blindness = dis.get("blindness", {}) or {}
                blindness_male = _to_int(blindness.get("male"), 0)
                blindness_female = _to_int(blindness.get("female"), 0)
                blindness_others = _to_int(blindness.get("others") or blindness.get("other"), 0)

                hearing = dis.get("hearing", {}) or {}
                hearing_male = _to_int(hearing.get("male"), 0)
                hearing_female = _to_int(hearing.get("female"), 0)
                hearing_others = _to_int(hearing.get("others") or hearing.get("other"), 0)

                locomotor = dis.get("locomotor", {}) or {}
                locomotor_male = _to_int(locomotor.get("male"), 0)
                locomotor_female = _to_int(locomotor.get("female"),0)
                locomotor_other = _to_int(locomotor.get("others") or locomotor.get("other"), 0)

                autism = dis.get("autism", {}) or {}
                autism_male = _to_int(autism.get("male"), 0)
                autism_female = _to_int(autism.get("female"),0)
                autism_other = _to_int(autism.get("others") or autism.get("other"), 0)

                other_disability = dis.get("other_disability", {}) or {}
                other_disability_male = _to_int(other_disability.get("male"), 0)
                other_disability_female = _to_int(other_disability.get("female"), 0)
                other_disability_other = _to_int(other_disability.get("others") or other_disability.get("other"), 0)


                defaults = {
                    "total_students": total_students,
                    "total_male": male,
                    "total_female": female,
                    "total_others": others,

                    "total_washrooms": total_washrooms,
                    "male_washrooms" : male_washrooms,
                    "female_washrooms": female_washrooms,

                    "open_male": open_male,
                    "open_female": open_female,
                    "open_others": open_others,

                    "obc_male": obc_male,
                    "obc_female": obc_female,
                    "obc_others": obc_others,

                    "sc_male": sc_male,
                    "sc_female": sc_female,
                    "sc_others": sc_others,

                    "st_male": st_male,
                    "st_female": st_female,
                    "st_others": st_others,

                    "nt_male": nt_male,
                    "nt_female":nt_female,
                    "nt_others": nt_others,

                    "vjnt_male": vjnt_male,
                    "vjnt_female": vjnt_female,
                    "vjnt_others": vjnt_others,

                    "ews_male": ews_male,
                    "ews_female": ews_female,
                    "ews_others": ews_others,

                    "hindu_male": hindu_male,
                    "hindu_female": hindu_female,
                    "hindu_others": hindu_others,

                    "muslim_male": muslim_male,
                    "muslim_female": muslim_female,
                    "muslim_others": muslim_others,
                    
                    "sikh_male": sikh_male,
                    "sikh_female": sikh_female,
                    "sikh_others": sikh_others,

                    "christian_male": christian_male,
                    "christian_female": christian_female,
                    "christian_others": christian_others,

                    "jain_male": jain_male,
                    "jain_female": jain_female, 
                    "jain_others": jain_others,

                    "buddhist_male": buddhist_male,
                    "buddhist_female": buddhist_female,
                    "buddhist_others": buddhist_others,

                    "other_religion_male": other_religion_male,
                    "other_religion_female": other_religion_female,
                    "other_religion_others": other_religion_others,

                    "no_disability_male": no_disability_male,
                    "no_disability_female": no_disability_female,
                    "no_disability_others": no_disability_others,

                    "low_vision_male": low_vision_male,
                    "low_vision_female": low_vision_female,
                    "low_vision_others": low_vision_others,

                    "blindness_male": blindness_male,
                    "blindness_female": blindness_female,
                    "blindness_others": blindness_others,

                    "hearing_male": hearing_male,
                    "hearing_female": hearing_female,
                    "hearing_others": hearing_others,

                    "locomotor_male": locomotor_male,
                    "locomotor_female": locomotor_female,
                    "locomotor_others": locomotor_other,

                    "autism_male": autism_male,
                    "autism_female": autism_female,
                    "autism_others": autism_other,

                    "other_disability_male": other_disability_male,
                    "other_disability_female": other_disability_female,
                    "other_disability_others": other_disability_other,
                }

                try:
                    client_ip = get_client_ip(request)

                    # Check only ACTIVE record (is_deleted = False)
                    existing = student_aggregate_master.objects.filter(
                        College=college,
                        Program=program_obj,
                        Academic_Year=academic_year,
                        is_deleted=False,
                    ).first()

                    if existing:
                        errors.append({
                            "program": program_name,
                            "error": "Record already exists for this college, program and academic year"
                        })
                        continue

                    # 🆕 No active record → create a completely NEW one
                    obj = student_aggregate_master.objects.create(
                        College=college,
                        Program=program_obj,
                        Academic_Year=academic_year,
                        is_deleted=False,
                        created_by=client_ip,
                        **defaults,
                    )

                    saved.append({"program": program_name, "id": obj.pk, "created": True})

                except Exception as e:
                    errors.append({"program": program_name, "error": f"DB error: {str(e)}"})
                    continue

        response_status = 200 if not errors else 207
        resp = {
            "status": response_status,
            "saved": saved,
            "errors": errors,
            "summary": {
                "created": sum(1 for s in saved if s.get("created")),
                "updated": 0,   # no updates here
                "failed": len(errors)
            }
        }
        return JsonResponse(resp)   


@ajax_login_required
def update_student_aggregate(request):
    if request.method == "POST":
        try:
            payload = json.loads(request.body)
        except json.JSONDecodeError:
            return HttpResponseBadRequest("Invalid JSON")

        college_code = payload.get('college_code')
        academic_year = payload.get('academic_year')
        records = payload.get('records', [])

        if not college_code or not academic_year:
            return JsonResponse({'status': 400, 'message': 'Missing required fields'})

        try:
            college = College.objects.get(College_Code=college_code, is_deleted=False)
        except College.DoesNotExist:
            return JsonResponse({'status': 404, 'message': 'College not found'})

        if not isinstance(records, dict) or len(records) == 0:
            return JsonResponse({"status": 400, "message": "No records provided"}, status=400)
        
        updated = []
        created = []
        errors = []

        with transaction.atomic():
            for program_name, data in records.items():
                program_obj = CollegeProgram.objects.filter(
                    College=college,
                    ProgramName=program_name,
                    is_deleted=False
                ).first()

                if not program_obj:
                    errors.append({
                        "program": program_name,
                        "error": f"Program '{program_name}' not found for college '{college_code}'"
                    })
                    continue

                # ---- parse numeric fields safely (same as add) ----
                total_students = _to_int(data.get("total_students"), 0)

                gender = data.get("gender", {}) or {}
                male = _to_int(gender.get("male"), 0)
                female = _to_int(gender.get("female"), 0)
                others = _to_int(gender.get("others") or gender.get("other"), 0)

                washrooms = data.get("washrooms", {}) or {}
                total_washrooms = _to_int(washrooms.get("total"), 0)
                male_washrooms = _to_int(washrooms.get("male"), 0)
                female_washrooms = _to_int(washrooms.get("female"), 0)

                category = data.get("category", {}) or {}
                open = category.get("open", {}) or {}
                open_male = _to_int(open.get("male"), 0)
                open_female = _to_int(open.get("female"), 0)
                open_others = _to_int(open.get("others") or open.get("other"), 0)

                obc = category.get("obc", {}) or {}
                obc_male = _to_int(obc.get("male"), 0)
                obc_female = _to_int(obc.get("female"), 0)
                obc_others = _to_int(obc.get("others") or obc.get("other"), 0)
                
                sc = category.get("sc", {}) or {}
                sc_male = _to_int(sc.get("male"), 0)
                sc_female = _to_int(sc.get("female"), 0)
                sc_others = _to_int(sc.get("others") or sc.get("other"), 0)

                st = category.get("st", {}) or {}
                st_male = _to_int(st.get("male"), 0)
                st_female = _to_int(st.get("female"), 0)
                st_others = _to_int(st.get("others") or st.get("other"), 0) 

                nt = category.get("nt", {}) or {}
                nt_male = _to_int(nt.get("male"), 0)
                nt_female = _to_int(nt.get("female"), 0)
                nt_others = _to_int(nt.get("others") or nt.get("other"), 0)

                vjnt = category.get("vjnt", {}) or {}
                vjnt_male = _to_int(vjnt.get("male"), 0)
                vjnt_female = _to_int(vjnt.get("female"), 0)
                vjnt_others = _to_int(vjnt.get("others") or vjnt.get("other"), 0)

                ews = category.get("ews", {}) or {}
                ews_male = _to_int(ews.get("male"), 0)
                ews_female = _to_int(ews.get("female"), 0)
                ews_others = _to_int(ews.get("others") or ews.get("other"), 0)


                religion = data.get("religion", {}) or {}

                hindu = religion.get("hindu", {}) or {}
                hindu_male = _to_int(hindu.get("male"), 0)
                hindu_female = _to_int(hindu.get("female"), 0)
                hindu_others = _to_int(hindu.get("others") or hindu.get("other"), 0)

                muslim = religion.get("muslim", {}) or {}
                muslim_male = _to_int(muslim.get("male"), 0)
                muslim_female = _to_int(muslim.get("female"), 0)
                muslim_others = _to_int(muslim.get("others") or muslim.get("other"), 0)

                sikh = religion.get("sikh", {}) or {}
                sikh_male = _to_int(sikh.get("male"), 0)
                sikh_female = _to_int(sikh.get("female"), 0)
                sikh_others = _to_int(sikh.get("others") or sikh.get("other"), 0)

                christian = religion.get("christian", {}) or {}
                christian_male = _to_int(christian.get("male"), 0)
                christian_female = _to_int(christian.get("female"), 0)
                christian_others = _to_int(christian.get("others") or christian.get("other"), 0)

                jain = religion.get("jain", {}) or {}
                jain_male = _to_int(jain.get("male"), 0)
                jain_female = _to_int(jain.get("female"), 0)
                jain_others = _to_int(jain.get("others") or jain.get("other"), 0)

                buddhist = religion.get("buddhist", {}) or {}
                buddhist_male = _to_int(buddhist.get("male"), 0)
                buddhist_female = _to_int(buddhist.get("female"), 0)
                buddhist_others = _to_int(buddhist.get("others") or buddhist.get("other"), 0)

                other_religion = religion.get("other_religion", {}) or {}
                other_religion_male = _to_int(other_religion.get("male"), 0)
                other_religion_female = _to_int(other_religion.get("female"), 0)
                other_religion_others = _to_int(other_religion.get("others") or other_religion.get("other"), 0)


                dis = data.get("disability", {}) or {}
                no_disability = dis.get("no_disability", {}) or {}
                no_disability_male = _to_int(no_disability.get("male"), 0)
                no_disability_female = _to_int(no_disability.get("female"), 0)
                no_disability_others = _to_int(no_disability.get("others") or no_disability.get("other"), 0)

                low_vision = dis.get("lowvision", {}) or {}
                low_vision_male = _to_int(low_vision.get("male"), 0)
                low_vision_female = _to_int(low_vision.get("female"), 0)
                low_vision_others = _to_int(low_vision.get("others") or low_vision.get("other"), 0)

                blindness = dis.get("blindness", {}) or {}
                blindness_male = _to_int(blindness.get("male"), 0)
                blindness_female = _to_int(blindness.get("female"), 0)
                blindness_others = _to_int(blindness.get("others") or blindness.get("other"), 0)

                hearing = dis.get("hearing", {}) or {}
                hearing_male = _to_int(hearing.get("male"), 0)
                hearing_female = _to_int(hearing.get("female"), 0)
                hearing_others = _to_int(hearing.get("others") or hearing.get("other"), 0)

                locomotor = dis.get("locomotor", {}) or {}
                locomotor_male = _to_int(locomotor.get("male"), 0)
                locomotor_female = _to_int(locomotor.get("female"),0)
                locomotor_others = _to_int(locomotor.get("others") or locomotor.get("other"), 0)

                autism = dis.get("autism", {}) or {}
                autism_male = _to_int(autism.get("male"), 0)
                autism_female = _to_int(autism.get("female"),0)
                autism_others = _to_int(autism.get("others") or autism.get("other"), 0)

                other_disability = dis.get("other_disability", {}) or {}
                other_disability_male = _to_int(other_disability.get("male"), 0)
                other_disability_female = _to_int(other_disability.get("female"), 0)
                other_disability_others = _to_int(other_disability.get("others") or other_disability.get("other"), 0)


                try:
                    client_ip = get_client_ip(request)

                    existing = student_aggregate_master.objects.filter(
                        College=college,
                        Program=program_obj,
                        Academic_Year=academic_year,
                        is_deleted=False,
                    ).first()

                    
                    if existing:
                        # Update fields
                        existing.total_students = total_students
                        existing.total_male = male
                        existing.total_female = female
                        existing.total_others = others

                        existing.total_washrooms = total_washrooms
                        existing.male_washrooms = male_washrooms
                        existing.female_washrooms = female_washrooms

                        existing.open_male = open_male
                        existing.open_female = open_female
                        existing.open_others = open_others

                        existing.obc_male = obc_male
                        existing.obc_female = obc_female
                        existing.obc_others = obc_others

                        existing.sc_male = sc_male
                        existing.sc_female = sc_female
                        existing.sc_others = sc_others

                        existing.st_male = st_male
                        existing.st_female = st_female
                        existing.st_others = st_others

                        existing.nt_male = nt_male
                        existing.nt_female = nt_female
                        existing.nt_others = nt_others

                        existing.vjnt_male = vjnt_male
                        existing.vjnt_female = vjnt_female
                        existing.vjnt_others = vjnt_others

                        existing.ews_male = ews_male
                        existing.ews_female = ews_female
                        existing.ews_others = ews_others

                        
                        existing.hindu_male = hindu_male
                        existing.hindu_female = hindu_female
                        existing.hindu_others = hindu_others

                        existing.muslim_male = muslim_male
                        existing.muslim_female = muslim_female
                        existing.muslim_others = muslim_others

                        existing.sikh_male = sikh_male
                        existing.sikh_female = sikh_female
                        existing.sikh_others = sikh_others

                        existing.christian_male = christian_male
                        existing.christian_female = christian_female
                        existing.christian_others = christian_others

                        existing.jain_male = jain_male
                        existing.jain_female = jain_female
                        existing.jain_others = jain_others

                        existing.buddhist_male = buddhist_male
                        existing.buddhist_female = buddhist_female
                        existing.buddhist_others = buddhist_others

                        existing.other_religion_male = other_religion_male
                        existing.other_religion_female = other_religion_female
                        existing.other_religion_others = other_religion_others

                        existing.no_disability_male = no_disability_male
                        existing.no_disability_female = no_disability_female
                        existing.no_disability_others = no_disability_others

                        existing.low_vision_male = low_vision_male
                        existing.low_vision_female = low_vision_female
                        existing.low_vision_others = low_vision_others

                        existing.blindness_male = blindness_male
                        existing.blindness_female = blindness_female
                        existing.blindness_others = blindness_others

                        existing.hearing_male = hearing_male
                        existing.hearing_female = hearing_female
                        existing.hearing_others = hearing_others

                        existing.locomotor_male = locomotor_male
                        existing.locomotor_female = locomotor_female
                        existing.locomotor_others = locomotor_others

                        existing.autism_male = autism_male
                        existing.autism_female = autism_female
                        existing.autism_others = autism_others

                        existing.other_disability_male = other_disability_male
                        existing.other_disability_female = other_disability_female
                        existing.other_disability_others = other_disability_others


                        existing.updated_by = client_ip
                        existing.save()

                        updated.append({"program": program_name, "id": existing.pk, "updated": True})
                    else:
                         # No existing row → CREATE a new one (for new program/year combo)
                        obj = student_aggregate_master.objects.create(
                            College=college,
                            Program=program_obj,
                            Academic_Year=academic_year,
                            is_deleted=False,
                            created_by=client_ip,

                            total_students=total_students,
                            total_male=male,
                            total_female=female,
                            total_others=others,

                            total_washrooms = total_washrooms,
                            male_washrooms = male_washrooms,
                            female_washrooms = female_washrooms,


                            open_male = open_male,
                            open_female = open_female,
                            open_others = open_others,

                            obc_male = obc_male,
                            obc_female = obc_female,
                            obc_others = obc_others,

                            sc_male = sc_male,
                            sc_female = sc_female,
                            sc_others = sc_others,

                            st_male = st_male,
                            st_female = st_female,
                            st_others = st_others,

                            nt_male = nt_male,
                            nt_female = nt_female,
                            nt_others = nt_others,

                            vjnt_male = vjnt_male,
                            vjnt_female = vjnt_female,
                            vjnt_others = vjnt_others,

                            ews_male = ews_male,
                            ews_female = ews_female,
                            ews_others = ews_others,

                            hindu_male = hindu_male,
                            hindu_female = hindu_female,
                            hindu_others = hindu_others,

                            muslim_male = muslim_male,
                            muslim_female = muslim_female,
                            muslim_others = muslim_others,

                            sikh_male = sikh_male,
                            sikh_female = sikh_female,
                            sikh_others = sikh_others,

                            christian_male = christian_male,
                            christian_female = christian_female,
                            christian_others = christian_others,

                            jain_male = jain_male,
                            jain_female = jain_female,
                            jain_others = jain_others,

                            buddhist_male = buddhist_male,
                            buddhist_female = buddhist_female,
                            buddhist_others = buddhist_others,

                            other_religion_male = other_religion_male,
                            other_religion_female = other_religion_female,
                            other_religion_others = other_religion_others,

                            no_disability_male = no_disability_male,
                            no_disability_female = no_disability_female,
                            no_disability_others = no_disability_others,

                            low_vision_male = low_vision_male,
                            low_vision_female = low_vision_female,
                            low_vision_others = low_vision_others,

                            blindness_male = blindness_male,
                            blindness_female = blindness_female,
                            blindness_others = blindness_others,

                            hearing_male = hearing_male,
                            hearing_female = hearing_female,
                            hearing_others = hearing_others,

                            locomotor_male = locomotor_male,
                            locomotor_female = locomotor_female,
                            locomotor_others = locomotor_others,

                            autism_male = autism_male,
                            autism_female = autism_female,
                            autism_others = autism_others,

                            other_disability_male = other_disability_male,
                            other_disability_female = other_disability_female,
                            other_disability_others = other_disability_others,


                        )
                        created.append({"program": program_name, "id": obj.pk, "created": True})

                except Exception as e:
                    errors.append({"program": program_name, "error": f"DB error: {str(e)}"})
                    continue

        response_status = 200 if not errors else 207
        resp = {
            "status": response_status,
            "updated": updated,
            "errors": errors,
            "summary": {
                "created": 0,
                "updated": sum(1 for u in updated if u.get("updated")),
                "failed": len(errors)
            }
        }
        return JsonResponse(resp)

    
@ajax_login_required
def get_student_records(request):
    try:
        draw = int(request.GET.get("draw", 1))
        start = int(request.GET.get("start", 0))
        length = int(request.GET.get("length", 10))
    except ValueError:
        return HttpResponseBadRequest("Invalid paging parameters")

    search_value = request.GET.get("search[value]", "").strip()
    order_col_index = request.GET.get("order[0][column]")
    order_dir = request.GET.get("order[0][dir]", "asc")
    year = request.GET.get("year")

    if not year:
        latest = academic_year.objects.order_by("-Academic_Year").first()
        year = latest.Academic_Year if latest else ""

    # 1) Base queryset: colleges that have student aggregates in this year
    base_qs = College.objects.filter(
        is_deleted=False,
        student_aggregates__Academic_Year=year,
        student_aggregates__is_deleted=False,
    ).distinct()

    # 2) Total before search
    records_total = base_qs.count()

    # 3) Apply global search across College + CollegeProgram
    if search_value:
        colleges_qs = base_qs.filter(
            Q(College_Code__icontains=search_value) |
            Q(College_Name__icontains=search_value) |
            Q(address__icontains=search_value) |
            Q(country__icontains=search_value) |
            Q(state__icontains=search_value) |
            Q(District__icontains=search_value) |
            Q(taluka__icontains=search_value) |
            Q(city__icontains=search_value) |
            Q(pincode__icontains=search_value) |
            Q(college_type__icontains=search_value) |
            Q(belongs_to__icontains=search_value) |
            Q(affiliated__icontains=search_value) |
            # from related CollegeProgram
            Q(college_programs__Discipline__icontains=search_value) |
            Q(college_programs__ProgramName__icontains=search_value)
        ).distinct()
    else:
        colleges_qs = base_qs

    # 4) Filtered count after search
    records_filtered = colleges_qs.count()

    # 5) Ordering
    order_map = {
        "1": "College_Code",
        "2": "College_Name",
    }

    if order_col_index == "4":
        # order by total students (aggregated per college)
        colleges_qs = colleges_qs.annotate(
            agg_total=Sum(
                "staff_aggregates__total_students",
                filter=Q(staff_aggregates__Academic_Year=year),
            )
        )
        field = "agg_total"
        if order_dir == "desc":
            field = "-" + field
        colleges_qs = colleges_qs.order_by(field, "College_Name")
    else:
        field = order_map.get(order_col_index, "College_Name")
        if order_dir == "desc":
            field = "-" + field
        colleges_qs = colleges_qs.order_by(field)

    # 6) Pagination
    colleges_page = colleges_qs[start : start + length]

    data = []

    for col in colleges_page:
        pc_qs = (
            student_aggregate_master.objects
            .filter(College=col, Academic_Year=year, is_deleted=False)
            .select_related("Program")
            .order_by("Program__Discipline", "Program__ProgramName")
        )

        discipline_map = {}
        total_students_for_college = 0

        for pc in pc_qs:
            prog_obj = pc.Program
            discipline = prog_obj.Discipline if prog_obj else "Unspecified"
            program_name = prog_obj.ProgramName if prog_obj else str(pc.Program_id)

            total_students_for_college += (pc.total_students or 0)

            entry = {
                "name": program_name,
                "total_students": pc.total_students or 0,
                "gender": {
                    "male": pc.total_male or 0,
                    "female": pc.total_female or 0,
                    "others": pc.total_others or 0,
                },
                "washrooms": {
                    "total_washrooms":pc.total_washrooms or 0,
                    "male_washrooms": pc.male_washrooms or 0,
                    "female_washrooms" : pc.female_washrooms or 0,
                },
                "category": {
                    "open": {
                        "male": pc.open_male or 0,
                        "female": pc.open_female or 0,
                        "others": pc.open_others or 0,
                    },
                    "obc": {
                        "male": pc.obc_male or 0,
                        "female": pc.obc_female or 0,
                        "others": pc.obc_others or 0,
                    },
                    "sc": {
                        "male": pc.sc_male or 0,
                        "female": pc.sc_female or 0,
                        "others": pc.sc_others or 0,
                    },
                    "st": {
                        "male": pc.st_male or 0,
                        "female": pc.st_female or 0,
                        "others": pc.st_others or 0,
                    },
                    "nt": {
                        "male": pc.nt_male or 0,
                        "female": pc.nt_female or 0,
                        "others": pc.nt_others or 0,
                    },
                    "vjnt": {
                        "male": pc.vjnt_male or 0,
                        "female": pc.vjnt_female or 0,
                        "others": pc.vjnt_others or 0,
                    },
                    "ews": {
                        "male": pc.ews_male or 0,
                        "female": pc.ews_female or 0,
                        "others": pc.ews_others or 0,
                    },
                },
                "religion": {
                    "hindu": {
                        "male": pc.hindu_male or 0,
                        "female": pc.hindu_female or 0,
                        "others": pc.hindu_others or 0,
                    },
                    "muslim": {
                        "male": pc.muslim_male or 0,
                        "female": pc.muslim_female or 0,
                        "others": pc.muslim_others or 0,
                    },
                    "sikh": {
                        "male": pc.sikh_male or 0,
                        "female": pc.sikh_female or 0,
                        "others": pc.sikh_others or 0,
                    },
                    "christian": {
                        "male": pc.christian_male or 0,
                        "female": pc.christian_female or 0,
                        "others": pc.christian_others or 0,
                    },
                    "jain": {
                        "male": pc.jain_male or 0,
                        "female": pc.jain_female or 0,
                        "others": pc.jain_others or 0,
                    },
                    "buddhist": {
                        "male": pc.buddhist_male or 0,
                        "female": pc.buddhist_female or 0,
                        "others": pc.buddhist_others or 0,
                    },
                    "other": {
                        "male": pc.other_religion_male or 0,
                        "female": pc.other_religion_female or 0,
                        "others": pc.other_religion_others or 0,
                    },
                },
                "disability": {
                    "no_disability": {
                        "male": pc.no_disability_male or 0,
                        "female": pc.no_disability_female or 0,
                        "others": pc.no_disability_others or 0,
                    },
                    "lowvision": {
                        "male": pc.low_vision_male or 0,
                        "female": pc.low_vision_female or 0,
                        "others": pc.low_vision_others or 0,
                    },
                    "blindness": {
                        "male": pc.blindness_male or 0,
                        "female": pc.blindness_female or 0,
                        "others": pc.blindness_others or 0,
                    },
                    "hearing": {
                        "male": pc.hearing_male or 0,
                        "female": pc.hearing_female or 0,
                        "others": pc.hearing_others or 0,
                    },
                    "locomotor": {
                        "male": pc.locomotor_male or 0,
                        "female": pc.locomotor_female or 0,
                        "others": pc.locomotor_others or 0,
                    },
                    "autism": {
                        "male": pc.autism_male or 0,
                        "female": pc.autism_female or 0,
                        "others": pc.autism_others or 0,
                    },
                    "other": {
                        "male": pc.other_disability_male or 0,
                        "female": pc.other_disability_female or 0,
                        "others": pc.other_disability_others or 0,
                    },
                },
            }

            discipline_map.setdefault(discipline, []).append(entry)

        grouped_list = []
        for disc in sorted(discipline_map.keys(), key=str.lower):
            grouped_list.append({
                "discipline": disc,
                "programs": sorted(discipline_map[disc], key=lambda x: x["name"].lower()),
            })

        data.append({
            "college_code": col.College_Code,
            "college_name": col.College_Name,
            "academic_year": year,
            "total_students": total_students_for_college,
            "programs": grouped_list,
        })

    return JsonResponse({
        "draw": draw,
        "recordsTotal": records_total,
        "recordsFiltered": records_filtered,
        "data": data,
    })


@ajax_login_required
def delete_student_record(request):
    if request.method == 'POST':
        college_code = request.POST.get('college_code')
        academic_year = request.POST.get('academic_year')

        try:
            college = College.objects.get(College_Code=college_code, is_deleted=False)
        except College.DoesNotExist:
            return JsonResponse({'status': 404, 'message': 'College not found'})

        student_aggregate_master.objects.filter(College=college, Academic_Year=academic_year, is_deleted=False).update(is_deleted=True)

        response_data = {
            'message': 'Student records deleted successfully',
            'status': 204
        }
        return JsonResponse(response_data)


def get_college_records(request):
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    total_count = College.objects.filter(is_deleted=False).count()

    program_prefetch = Prefetch(
        'college_programs',
        queryset=CollegeProgram.objects.filter(is_deleted=False),
        to_attr='program_list'
    )
    qs = College.objects.filter(is_deleted=False)

    if search_value:
        qs = qs.filter(
            Q(College_Code__icontains=search_value) |
            Q(College_Name__icontains=search_value) |
            Q(address__icontains=search_value) |
            Q(country__icontains=search_value) |
            Q(state__icontains=search_value) |
            Q(District__icontains=search_value) |
            Q(taluka__icontains=search_value) |
            Q(city__icontains=search_value) |
            Q(pincode__icontains=search_value) |
            Q(college_type__icontains=search_value) |
            Q(belongs_to__icontains=search_value) |
            Q(affiliated__icontains=search_value) |
            Q(college_programs__Discipline__icontains=search_value) |
            Q(college_programs__ProgramName__icontains=search_value)
        ).distinct()

    qs = qs.prefetch_related(program_prefetch)
    filtered_count = qs.count()

    # ordering (optional - keep simple)
    column_index = int(request.GET.get('order[0][column]', 2))
    direction = request.GET.get('order[0][dir]', 'asc')
    column_map = {
        1: 'College_Code',
        2: 'College_Name',
        3: 'state',
        4: 'District',
        5: 'city'
    }
    order_field = column_map.get(column_index, 'College_Name')
    if direction == 'desc':
        order_field = '-' + order_field
    qs = qs.order_by(order_field)

    # pagination
    qs_page = qs[start:start + length]

    # Build response rows as objects (so JS can access by property)
    data = []
    for college in qs_page:
        # build programs grouped by discipline
        programs_map = {}
        for p in getattr(college, 'program_list', []):
            programs_map.setdefault(p.Discipline or "Other", []).append(p.ProgramName)

        row = {
            "college_code": college.College_Code,
            "college_name": college.College_Name,
            "address": college.address,
            "country": college.country if hasattr(college, "country") else "",
            "state": college.state,
            "district": college.District,
            "taluka": college.taluka,
            "city": college.city,
            "pincode": college.pincode,
            "college_type": college.college_type,
            "belongs_to": college.belongs_to,
            "affiliated": college.affiliated if hasattr(college, "affiliated") else "",
            "programs": programs_map,
            "id": college.id
        }
        data.append(row)

    return JsonResponse({
        "draw": draw,
        "recordsTotal": total_count,
        "recordsFiltered": filtered_count,
        "data": data
    })

# Map DataTables column index -> model field name (adjust if your model fields differ)
# Indices follow your JS columns: 0 = toggle, 1 = college_code, 2 = college_name, ...
COLUMN_INDEX_TO_FIELD = {
    1: 'College_Code',
    2: 'College_Name',
    3: 'state',
    4: 'District',
    5: 'taluka',
    6: 'college_type',
    7: 'belongs_to',
    8: 'country',
    9: 'affiliated',
}


def parse_payload(request):
    """
    Try to parse JSON body first; fall back to POST['params'] or POST['data'].
    Returns dict or raises ValueError.
    """
    try:
        body = request.body.decode('utf-8')
        if body:
            return json.loads(body)
    except Exception:
        pass

    raw = request.POST.get('params') or request.POST.get('data')
    if not raw:
        raise ValueError("Invalid request payload")
    try:
        return json.loads(raw)
    except Exception:
        raise ValueError("Unable to parse params JSON")


@ajax_login_required
def export_colleges_excel(request):

    if request.method != "POST":
        return HttpResponseBadRequest("Only POST requests allowed.")
    try:
        payload = parse_payload(request)
    except ValueError as e:
        return HttpResponseBadRequest(str(e))

    search_text = (payload.get('search') or '').strip()
    order_instructions = payload.get('order', [])
    extra_filters = payload.get('extra_filters', {}) or {}

    # Base queryset - replicate the same base as your original view
    qs = College.objects.filter(is_deleted=False).prefetch_related("college_programs")

    # Apply extra_filters if you use any (example placeholder)
    # if extra_filters.get('state'):
    #     qs = qs.filter(state=extra_filters['state'])

    # Global search across relevant fields (keeps same behavior as DataTables search)
    if search_text:
        qs = qs.filter(
            Q(College_Code__icontains=search_text) |
            Q(College_Name__icontains=search_text) |
            Q(address__icontains=search_text) |
            Q(country__icontains=search_text) |
            Q(state__icontains=search_text) |
            Q(District__icontains=search_text) |
            Q(taluka__icontains=search_text) |
            Q(city__icontains=search_text) |
            Q(pincode__icontains=search_text) |
            Q(college_type__icontains=search_text) |
            Q(belongs_to__icontains=search_text) |
            Q(affiliated__icontains=search_text) |
            Q(college_programs__Discipline__icontains=search_text) |
            Q(college_programs__ProgramName__icontains=search_text)
        )

    # Apply ordering mapped from DataTables indices to model fields (if provided)
    if order_instructions and isinstance(order_instructions, list):
        order_by = []
        for ord_pair in order_instructions:
            try:
                col_index = int(ord_pair[0])
                direction = ord_pair[1].lower() if len(ord_pair) > 1 else 'asc'
                field = COLUMN_INDEX_TO_FIELD.get(col_index)
                if field:
                    prefix = '-' if direction == 'desc' else ''
                    order_by.append(prefix + field)
            except Exception:
                continue
        if order_by:
            qs = qs.order_by(*order_by)

    qs = qs.distinct()

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "College Records"

    # ========= HEADER ==========
    headers = [
        "College Code", "College Name", "Address", "Pincode", "Country",
        "State", "District", "Taluka", "City",
        "College Type", "Belongs To", "Affiliated To",
        # removed college-level washrooms here
        "Discipline", "Program"
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="006699", end_color="006699", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_num = 2  # start from row 2

    # Iterate over colleges (qs already filtered & ordered)
    for college in qs:
        # Build program queryset for this college
        college_program_qs = college.college_programs.filter(is_deleted=False)

        # If user used the datatable searchbar, interpret search_text to restrict programs:
        # - include programs whose Discipline or ProgramName contain the search_text
        # - if none match for this college, skip the college entirely (user searched specifically)
        if search_text:
            allowed_cps = list(college_program_qs.filter(
                Q(Discipline__icontains=search_text) | Q(ProgramName__icontains=search_text)
            ))
            # If the global search matched the college-level fields (eg. college name) but no programs
            # match the search_text, user requested filtering by discipline/program→skip this college.
            if not allowed_cps:
                # If you prefer to still include colleges matched by college fields even when no programs match,
                # comment out this continue and set allowed_cps = list(college_program_qs)
                continue
        else:
            # No search_text → include all programs
            allowed_cps = list(college_program_qs)

        # If still no programs (college has none), create a single placeholder row
        if not allowed_cps:
            allowed_cps = [None]

        # Group allowed programs under each discipline
        discipline_map = {}
        for cp in allowed_cps:
            if cp is None:
                discipline_map.setdefault("No Discipline", []).append(None)
            else:
                disc = getattr(cp, 'Discipline', None) or 'No Discipline'
                prog = getattr(cp, 'ProgramName', None) or 'No Program'
                discipline_map.setdefault(disc, []).append(prog)

        discipline_list = list(discipline_map.items())

        # Total rows needed = sum of all programs across all disciplines
        total_program_rows = sum(len(programs) for _, programs in discipline_list)
        if total_program_rows == 0:
            total_program_rows = 1  # ensure at least one row per college

        start_row = row_num
        end_row = row_num + total_program_rows - 1

        # ========== MERGE ALL COLLEGE INFO CELLS ==========
        college_fields = [
            getattr(college, 'College_Code', '') or getattr(college, 'college_code', ''),
            getattr(college, 'College_Name', '') or getattr(college, 'college_name', ''),
            getattr(college, 'address', '') or '',
            getattr(college, 'pincode', '') or '',
            getattr(college, 'country', '') or '',
            getattr(college, 'state', '') or '',
            getattr(college, 'District', '') or getattr(college, 'district', ''),
            getattr(college, 'taluka', '') or '',
            getattr(college, 'city', '') or '',
            getattr(college, 'college_type', '') or '',
            getattr(college, 'belongs_to', '') or '',
            getattr(college, 'affiliated', '') or '',
        ]

        for col_index, value in enumerate(college_fields, start=1):
            ws.merge_cells(
                start_row=start_row, start_column=col_index,
                end_row=end_row, end_column=col_index
            )
            cell = ws.cell(row=start_row, column=col_index, value=value)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        # ========== WRITE DISCIPLINE + PROGRAM CELLS ==========
        current_row = row_num

        for discipline, programs in discipline_list:
            discipline_rowspan = len(programs) if programs else 1
            discipline_start_row = current_row
            discipline_end_row = current_row + discipline_rowspan - 1

            # Merge discipline cell (column 13)
            ws.merge_cells(
                start_row=discipline_start_row, start_column=13,
                end_row=discipline_end_row, end_column=13
            )
            disc_cell = ws.cell(row=discipline_start_row, column=13, value=discipline)
            disc_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

            # Write each program in its own row (column 14)
            if programs:
                for p in programs:
                    prog_val = p or ''
                    prog_cell = ws.cell(current_row, 14, prog_val)
                    prog_cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                    current_row += 1
            else:
                prog_cell = ws.cell(current_row, 14, '')
                prog_cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                current_row += 1

        # Move pointer after writing all rows for this college
        row_num = end_row + 1

    # Auto column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for cell in ws[column_letter]:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[column_letter].width = min(max_length + 5, 60)

    # ========= RETURN FILE ==========
    date_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"College_Detailed_Report_{date_str}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@ajax_login_required
def export_student_excel(request):
    if request.method != "POST":
        return HttpResponseBadRequest("Only POST allowed")

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return HttpResponseBadRequest("Invalid JSON payload")

    year = payload.get("year")
    if not year:
        return HttpResponseBadRequest("Missing academic year")

    global_search = (payload.get("search") or "").strip()
    order_instructions = payload.get("order", []) or []

    # -----------------------
    # Base queryset
    # -----------------------
    qs = College.objects.filter(
        is_deleted=False,
        student_aggregates__Academic_Year=year,
        student_aggregates__is_deleted=False
    ).prefetch_related("college_programs", "student_aggregates").distinct()

    # -----------------------
    # Global Search - keep same fields as DataTable
    # -----------------------
    if global_search:
        qs = qs.filter(
            Q(College_Name__icontains=global_search)
            | Q(College_Code__icontains=global_search)
            | Q(address__icontains=global_search)
            | Q(country__icontains=global_search)
            | Q(state__icontains=global_search)
            | Q(District__icontains=global_search)
            | Q(taluka__icontains=global_search)
            | Q(city__icontains=global_search)
            | Q(pincode__icontains=global_search)
            | Q(college_programs__ProgramName__icontains=global_search)
            | Q(college_programs__Discipline__icontains=global_search)
        ).distinct()

    # -----------------------
    # Ordering (DataTables-style)
    # -----------------------
    COLUMN_INDEX_TO_FIELD = {
        1: "College_Code",
        2: "College_Name",
        3: "student_aggregates__Academic_Year",
        4: "student_aggregates__total_students",
    }

    order_by = []
    try:
        for pair in order_instructions:
            if isinstance(pair, (list, tuple)) and len(pair) >= 2:
                cidx = int(pair[0])
                direction = (pair[1] or "asc").lower()
            elif isinstance(pair, dict):
                cidx = int(pair.get("column") or pair.get("col") or pair.get("0", 0))
                direction = (pair.get("dir") or "asc").lower()
            else:
                continue
            field = COLUMN_INDEX_TO_FIELD.get(cidx)
            if field:
                order_by.append(("-" if direction == "desc" else "") + field)
    except Exception:
        pass

    if order_by:
        qs = qs.order_by(*order_by)

    # -----------------------
    # Build Excel File
    # -----------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Records"

    # ---------- HEADERS ----------
    headers = [
        "College Code", "College Name", "Address", "Pincode", "Country",
        "State", "District", "Taluka", "City",
        "College Type", "Belongs To", "Affiliated To",

        "Discipline", "Program",

        # Student data columns start here
        "Total Students Washrooms",
        "Male Students washrooms",
        "Female Students washrooms",

        "Total Students",
        "Total Male", "Total Female", "Total Others",

        # caste
        "OPEN Male", "OPEN Female", "OPEN Others",
        "OBC Male", "OBC Female", "OBC Others",
        "SC Male", "SC Female", "SC Others",
        "ST Male", "ST Female", "ST Others",
        "NT Male", "NT Female", "NT Others",
        "VJNT Male", "VJNT Female", "VJNT Others",
        "EWS Male", "EWS Female", "EWS Others",

        # religion
        "Hindu Male", "Hindu Female", "Hindu Others",
        "Muslim Male", "Muslim Female", "Muslim Others",
        "Sikh Male", "Sikh Female", "Sikh Others",
        "Christian Male", "Christian Female", "Christian Others",
        "Jain Male", "Jain Female", "Jain Others",
        "Buddhist Male", "Buddhist Female", "Buddhist Others",
        "Other Religion Male", "Other Religion Female", "Other Religion Others",

        # disability
        "No Disability Male", "No Disability Female", "No Disability Others",
        "Low Vision Male", "Low Vision Female", "Low Vision Others",
        "Blindness Male", "Blindness Female", "Blindness Others",
        "Hearing Impaired Male", "Hearing Impaired Female", "Hearing Impaired Others",
        "Locomotor Disability Male", "Locomotor Disability Female", "Locomotor Disability Others",
        "Autism Male", "Autism Female", "Autism Others",
        "Other Disability Male", "Other Disability Female", "Other Disability Others",
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="006699", fill_type="solid")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    student_fields = [
        "total_washrooms", "male_washrooms", "female_washrooms",
        "total_students",
        "total_male", "total_female", "total_others",
        "open_male", "open_female", "open_others",
        "obc_male", "obc_female", "obc_others",
        "sc_male", "sc_female", "sc_others",
        "st_male", "st_female", "st_others",
        "nt_male", "nt_female", "nt_others",
        "vjnt_male", "vjnt_female", "vjnt_others",
        "ews_male", "ews_female", "ews_others",
        "hindu_male", "hindu_female", "hindu_others",
        "muslim_male", "muslim_female", "muslim_others",
        "sikh_male", "sikh_female", "sikh_others",
        "christian_male", "christian_female", "christian_others",
        "jain_male", "jain_female", "jain_others",
        "buddhist_male", "buddhist_female", "buddhist_others",
        "other_religion_male", "other_religion_female", "other_religion_others",
        "no_disability_male", "no_disability_female", "no_disability_others",
        "low_vision_male", "low_vision_female", "low_vision_others",
        "blindness_male", "blindness_female", "blindness_others",
        "hearing_male", "hearing_female", "hearing_others",
        "locomotor_male", "locomotor_female", "locomotor_others",
        "autism_male", "autism_female", "autism_others",
        "other_disability_male", "other_disability_female", "other_disability_others",
    ]

    overall_agg = {f: 0 for f in student_fields}

    row_num = 2

    # -----------------------
    # MAIN LOOP — includes discipline/program filtering logic
    # -----------------------
    for college in qs:

        # 1) Find only the year aggregates for this college
        year_records = {
            r.Program_id: r
            for r in college.student_aggregates.filter(Academic_Year=year, is_deleted=False)
        }

        # 2) Determine allowed college programs based on global_search
        master_programs = college.college_programs.filter(is_deleted=False)

        if global_search:
            # only allow programs whose discipline or program name match search
            allowed_programs = list(master_programs.filter(
                Q(Discipline__icontains=global_search) |
                Q(ProgramName__icontains=global_search)
            ))

            # If the global search matched the college-level fields but no programs match the search_text,
            # we skip the college (user is searching specifically for program/discipline).
            if not allowed_programs:
                continue
        else:
            # no search filtering → include all programs
            allowed_programs = list(master_programs)

        # Guard: ensure at least one placeholder row when there are no programs
        if not allowed_programs:
            allowed_programs = [None]

        # 3) Build discipline → program_list mapping from ONLY allowed programs
        discipline_map = {}
        for cp in allowed_programs:
            if cp is None:
                discipline_map.setdefault("No Discipline", []).append(None)
            else:
                discipline_map.setdefault(cp.Discipline or "Unspecified", []).append(cp)

        discipline_list = list(discipline_map.items())

        # 4) Setup row merging
        total_program_rows = sum(len(v) if v else 1 for _, v in discipline_list)
        if total_program_rows <= 0:
            total_program_rows = 1

        start_row = row_num
        end_row = row_num + total_program_rows - 1

        # 5) Write college info (merged)
        college_fields = [
            college.College_Code,
            college.College_Name,
            college.address or "",
            college.pincode or "",
            college.country or "",
            college.state or "",
            college.District or "",
            college.taluka or "",
            college.city or "",
            college.college_type or "",
            college.belongs_to or "",
            college.affiliated or "",
        ]

        for ci, val in enumerate(college_fields, start=1):
            ws.merge_cells(
                start_row=start_row,
                start_column=ci,
                end_row=end_row,
                end_column=ci
            )
            cell = ws.cell(row=start_row, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        # 6) Write programs
        current_row = row_num
        for discipline, program_objs in discipline_list:
            # ensure program_objs is non-empty
            if not program_objs:
                program_objs = [None]

            # safe merge for discipline column (column 13)
            ws.merge_cells(
                start_row=current_row,
                start_column=13,
                end_row=current_row + len(program_objs) - 1,
                end_column=13
            )
            ws.cell(row=current_row, column=13, value=discipline).alignment = Alignment(
                vertical="center", horizontal="center", wrap_text=True
            )

            for cp in program_objs:
                if cp:
                    ws.cell(row=current_row, column=14, value=cp.ProgramName)
                    rec = year_records.get(cp.id)
                else:
                    ws.cell(row=current_row, column=14, value="No Program")
                    rec = None

                # student data columns start at 15
                for i, field in enumerate(student_fields):
                    val = getattr(rec, field, 0) if rec else 0
                    ws.cell(row=current_row, column=15 + i, value=val)
                    overall_agg[field] += val if val else 0

                current_row += 1

        row_num = end_row + 1

    # -----------------------
    # FINAL AGGREGATE ROW
    # -----------------------
    agg_row = row_num
    ws.merge_cells(start_row=agg_row, start_column=1, end_row=agg_row, end_column=12)
    label_cell = ws.cell(row=agg_row, column=1, value=f"Aggregate Values - {year}")
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Key totals (student area starts at column 15)
    ws.cell(row=agg_row, column=15, value=overall_agg.get("total_washrooms", 0))
    ws.cell(row=agg_row, column=16, value=overall_agg.get("male_washrooms", 0))
    ws.cell(row=agg_row, column=17, value=overall_agg.get("female_washrooms", 0))
    ws.cell(row=agg_row, column=18, value=overall_agg.get("total_students", 0))

    for col_idx in (15, 16, 17, 18):
        c = ws.cell(row=agg_row, column=col_idx)
        c.font = Font(bold=True, color="CC6600")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, field in enumerate(student_fields):
        tot = overall_agg.get(field, 0)
        c = ws.cell(row=agg_row, column=15 + i, value=tot)
        c.font = Font(bold=True, color="CC6600")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-width
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
        ws.column_dimensions[letter].width = min(max_len + 5, 60)

    # Return XLSX
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    date_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Student_Report_{year}_{date_str}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@ajax_login_required
def export_dashboard_excel(request):
    if request.method != "POST":
       return JsonResponse(
            {"message": "Only POST allowed"},
            status=400
        )
    # --- Academic year: REQUIRED from POST ---
    year = (request.POST.get("year") or "").strip()
    if not year:
         return JsonResponse(
            {"message": "Missing academic year"},
            status=400
        )

    def non_empty_list(key):
        return [v for v in request.POST.getlist(key) if v is not None and v != ""]

    college_codes   = non_empty_list("CollegeCode[]")
    college_names   = non_empty_list("CollegeName[]")
    districts       = non_empty_list("District[]")
    talukas         = non_empty_list("Taluka[]")
    college_types   = non_empty_list("CollegeType[]")
    belongs_to_list = non_empty_list("BelongsTo[]")
    disciplines     = non_empty_list("Discipline[]")
    programs        = non_empty_list("Programs[]")

    # --------- BASE QS: all active colleges ----------
    qs = College.objects.filter(is_deleted=False)

    # 🔒 Restrict by logged-in user's college (if not superuser)
    user = request.user
    if not user.is_superuser:
        profile = UserCollege.objects.filter(user=user).first()
        if not profile or not profile.college or profile.college.is_deleted:
            # normal user but no valid college assigned
           return JsonResponse(
                {"message": "No college assigned to this user."},
                status=400
            )
        qs = qs.filter(id=profile.college.id)

    # prefetch AFTER user restriction
    qs = qs.prefetch_related(
        "college_programs", "student_aggregates", "staff_aggregates"
    )

    # --------- Apply filters from request ----------
    if college_codes:
        qs = qs.filter(College_Code__in=college_codes)
    if college_names:
        qs = qs.filter(College_Name__in=college_names)
    if districts:
        qs = qs.filter(District__in=districts)
    if talukas:
        qs = qs.filter(taluka__in=talukas)
    if college_types:
        qs = qs.filter(college_type__in=college_types)
    if belongs_to_list:
        qs = qs.filter(belongs_to__in=belongs_to_list)
    # Keep initial filtering if user provided disciplines/programs (this reduces the college list)
    if disciplines:
        qs = qs.filter(college_programs__Discipline__in=disciplines)
    if programs:
        qs = qs.filter(college_programs__ProgramName__in=programs)

    qs = qs.distinct().order_by("College_Name")

    # ---------- Build workbook with two sheets ----------
    wb = Workbook()
    # default active sheet -> students
    ws_students = wb.active
    ws_students.title = "Students"

    # create staff sheet
    ws_staff = wb.create_sheet(title="Staff")

    # Common meta headers
    meta_headers = [
        "College Code", "College Name", "Address", "Pincode", "Country",
        "State", "District", "Taluka", "City",
        "College Type", "Belongs To", "Affiliated To",
        "Discipline", "Program",
    ]

    # STUDENT headers (with "(Students)" style)
    student_headers = [
        "Total Students Washrooms", "Male Students Washrooms", "Female Students Washrooms",
        "Total Students", "Total Male (Students)", "Total Female (Students)", "Total Others (Students)",
        # caste
        "OPEN Male (Students)", "OPEN Female (Students)", "OPEN Others (Students)",
        "OBC Male (Students)", "OBC Female (Students)", "OBC Others (Students)",
        "SC Male (Students)", "SC Female (Students)", "SC Others (Students)",
        "ST Male (Students)", "ST Female (Students)", "ST Others (Students)",
        "NT Male (Students)", "NT Female (Students)", "NT Others (Students)",
        "VJNT Male (Students)", "VJNT Female (Students)", "VJNT Others (Students)",
        "EWS Male (Students)", "EWS Female (Students)", "EWS Others (Students)",
        # religion
        "Hindu Male (Students)", "Hindu Female (Students)", "Hindu Others (Students)",
        "Muslim Male (Students)", "Muslim Female (Students)", "Muslim Others (Students)",
        "Sikh Male (Students)", "Sikh Female (Students)", "Sikh Others (Students)",
        "Christian Male (Students)", "Christian Female (Students)", "Christian Others (Students)",
        "Jain Male (Students)", "Jain Female (Students)", "Jain Others (Students)",
        "Buddhist Male (Students)", "Buddhist Female (Students)", "Buddhist Others (Students)",
        "Other Religion Male (Students)", "Other Religion Female (Students)", "Other Religion Others (Students)",
        # disability
        "No Disability Male (Students)", "No Disability Female (Students)", "No Disability Others (Students)",
        "Low Vision Male (Students)", "Low Vision Female (Students)", "Low Vision Others (Students)",
        "Blindness Male (Students)", "Blindness Female (Students)", "Blindness Others (Students)",
        "Hearing Impaired Male (Students)", "Hearing Impaired Female (Students)", "Hearing Impaired Others (Students)",
        "Locomotor Disability Male (Students)", "Locomotor Disability Female (Students)", "Locomotor Disability Others (Students)",
        "Autism Male (Students)", "Autism Female (Students)", "Autism Others (Students)",
        "Other Disability Male (Students)", "Other Disability Female (Students)", "Other Disability Others (Students)",
    ]

    # STAFF headers (parallel style)
    staff_headers = [
        "Total Staff Washrooms", "Male Staff Washrooms", "Female Staff Washrooms",
        "Total Staff", "Total Male (Staff)", "Total Female (Staff)", "Total Others (Staff)",
        # caste
        "OPEN Male (Staff)", "OPEN Female (Staff)", "OPEN Others (Staff)",
        "OBC Male (Staff)", "OBC Female (Staff)", "OBC Others (Staff)",
        "SC Male (Staff)", "SC Female (Staff)", "SC Others (Staff)",
        "ST Male (Staff)", "ST Female (Staff)", "ST Others (Staff)",
        "NT Male (Staff)", "NT Female (Staff)", "NT Others (Staff)",
        "VJNT Male (Staff)", "VJNT Female (Staff)", "VJNT Others (Staff)",
        "EWS Male (Staff)", "EWS Female (Staff)", "EWS Others (Staff)",
        # religion
        "Hindu Male (Staff)", "Hindu Female (Staff)", "Hindu Others (Staff)",
        "Muslim Male (Staff)", "Muslim Female (Staff)", "Muslim Others (Staff)",
        "Sikh Male (Staff)", "Sikh Female (Staff)", "Sikh Others (Staff)",
        "Christian Male (Staff)", "Christian Female (Staff)", "Christian Others (Staff)",
        "Jain Male (Staff)", "Jain Female (Staff)", "Jain Others (Staff)",
        "Buddhist Male (Staff)", "Buddhist Female (Staff)", "Buddhist Others (Staff)",
        "Other Religion Male (Staff)", "Other Religion Female (Staff)", "Other Religion Others (Staff)",
        # disability
        "No Disability Male (Staff)", "No Disability Female (Staff)", "No Disability Others (Staff)",
        "Low Vision Male (Staff)", "Low Vision Female (Staff)", "Low Vision Others (Staff)",
        "Blindness Male (Staff)", "Blindness Female (Staff)", "Blindness Others (Staff)",
        "Hearing Impaired Male (Staff)", "Hearing Impaired Female (Staff)", "Hearing Impaired Others (Staff)",
        "Locomotor Disability Male (Staff)", "Locomotor Disability Female (Staff)", "Locomotor Disability Others (Staff)",
        "Autism Male (Staff)", "Autism Female (Staff)", "Autism Others (Staff)",
        "Other Disability Male (Staff)", "Other Disability Female (Staff)", "Other Disability Others (Staff)",
    ]

    # Write headers on both sheets
    headers_students = meta_headers + student_headers
    headers_staff = meta_headers + staff_headers

    ws_students.append(headers_students)
    ws_staff.append(headers_staff)

    # style header for both sheets
    header_fill = PatternFill(start_color="006699", fill_type="solid")
    for ws, headers in ((ws_students, headers_students), (ws_staff, headers_staff)):
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # field lists for reading from model instances
    student_fields = [
        "total_washrooms", "male_washrooms", "female_washrooms",
        "total_students", "total_male", "total_female", "total_others",
        # caste
        "open_male", "open_female", "open_others",
        "obc_male", "obc_female", "obc_others",
        "sc_male", "sc_female", "sc_others",
        "st_male", "st_female", "st_others",
        "nt_male", "nt_female", "nt_others",
        "vjnt_male", "vjnt_female", "vjnt_others",
        "ews_male", "ews_female", "ews_others",
        # religion
        "hindu_male", "hindu_female", "hindu_others",
        "muslim_male", "muslim_female", "muslim_others",
        "sikh_male", "sikh_female", "sikh_others",
        "christian_male", "christian_female", "christian_others",
        "jain_male", "jain_female", "jain_others",
        "buddhist_male", "buddhist_female", "buddhist_others",
        "other_religion_male", "other_religion_female", "other_religion_others",
        # disability
        "no_disability_male", "no_disability_female", "no_disability_others",
        "low_vision_male", "low_vision_female", "low_vision_others",
        "blindness_male", "blindness_female", "blindness_others",
        "hearing_male", "hearing_female", "hearing_others",
        "locomotor_male", "locomotor_female", "locomotor_others",
        "autism_male", "autism_female", "autism_others",
        "other_disability_male", "other_disability_female", "other_disability_others",
    ]

    staff_fields = [
        "total_staff_washrooms", "male_staff_washrooms", "female_staff_washrooms",
        "total_staff", "total_male", "total_female", "total_others",
        # caste
        "open_male", "open_female", "open_others",
        "obc_male", "obc_female", "obc_others",
        "sc_male", "sc_female", "sc_others",
        "st_male", "st_female", "st_others",
        "nt_male", "nt_female", "nt_others",
        "vjnt_male", "vjnt_female", "vjnt_others",
        "ews_male", "ews_female", "ews_others",
        # religion
        "hindu_male", "hindu_female", "hindu_others",
        "muslim_male", "muslim_female", "muslim_others",
        "sikh_male", "sikh_female", "sikh_others",
        "christian_male", "christian_female", "christian_others",
        "jain_male", "jain_female", "jain_others",
        "buddhist_male", "buddhist_female", "buddhist_others",
        "other_religion_male", "other_religion_female", "other_religion_others",
        # disability
        "no_disability_male", "no_disability_female", "no_disability_others",
        "low_vision_male", "low_vision_female", "low_vision_others",
        "blindness_male", "blindness_female", "blindness_others",
        "hearing_male", "hearing_female", "hearing_others",
        "locomotor_male", "locomotor_female", "locomotor_others",
        "autism_male", "autism_female", "autism_others",
        "other_disability_male", "other_disability_female", "other_disability_others",
    ]

    overall_student_agg = {f: 0 for f in student_fields}
    overall_staff_agg = {f: 0 for f in staff_fields}

    # Starting row indices for each sheet
    row_students = 2
    row_staff = 2

    # ---------- Iterate colleges (with restricted programs per filters) ----------
    for college in qs:
        # Determine allowed programs for this college based on filters
        college_program_qs = college.college_programs.filter(is_deleted=False)

        if programs:
            allowed_cps = list(college_program_qs.filter(ProgramName__in=programs))
        elif disciplines:
            allowed_cps = list(college_program_qs.filter(Discipline__in=disciplines))
        else:
            allowed_cps = list(college_program_qs)

        if not allowed_cps:
            allowed_cps = [None]

        allowed_prog_ids = [cp.id for cp in allowed_cps if cp is not None]

        student_year_records = {
            rec.Program_id: rec
            for rec in college.student_aggregates.filter(
                Academic_Year=year, is_deleted=False,
                **({"Program__in": allowed_prog_ids} if allowed_prog_ids else {})
            )
        }
        staff_year_records = {
            rec.Program_id: rec
            for rec in college.staff_aggregates.filter(
                Academic_Year=year, is_deleted=False,
                **({"Program__in": allowed_prog_ids} if allowed_prog_ids else {})
            )
        }

        discipline_map = {}
        for cp in allowed_cps:
            if cp is None:
                discipline_map.setdefault("No Discipline", []).append(None)
            else:
                discipline_map.setdefault(cp.Discipline or "Unspecified", []).append(cp)

        total_program_rows = sum(len(plist) if plist else 1 for _, plist in discipline_map.items())
        if total_program_rows <= 0:
            total_program_rows = 1

        # --- STUDENT sheet: merge meta across rows ---
        start_row = row_students
        end_row = row_students + total_program_rows - 1
        college_fields = [
            college.College_Code,
            college.College_Name,
            college.address or "",
            college.pincode or "",
            college.country or "",
            college.state or "",
            college.District or "",
            college.taluka or "",
            college.city or "",
            college.college_type or "",
            college.belongs_to or "",
            college.affiliated or "",
        ]
        for ci, val in enumerate(college_fields, start=1):
            ws_students.merge_cells(start_row=start_row, start_column=ci, end_row=end_row, end_column=ci)
            cell = ws_students.cell(row=start_row, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        # --- STAFF sheet: merge meta across rows (same)
        start_row_s = row_staff
        end_row_s = row_staff + total_program_rows - 1
        for ci, val in enumerate(college_fields, start=1):
            ws_staff.merge_cells(start_row=start_row_s, start_column=ci, end_row=end_row_s, end_column=ci)
            cell = ws_staff.cell(row=start_row_s, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        cur_row_s = row_students
        cur_row_f = row_staff

        for discipline, progs in discipline_map.items():
            progs_list = progs if progs else [None]

            ws_students.merge_cells(
                start_row=cur_row_s, start_column=13,
                end_row=cur_row_s + len(progs_list) - 1, end_column=13
            )
            ws_students.cell(cur_row_s, 13, discipline).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

            ws_staff.merge_cells(
                start_row=cur_row_f, start_column=13,
                end_row=cur_row_f + len(progs_list) - 1, end_column=13
            )
            ws_staff.cell(cur_row_f, 13, discipline).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

            for prog in progs_list:
                if prog:
                    ws_students.cell(cur_row_s, 14, prog.ProgramName)
                    ws_staff.cell(cur_row_f, 14, prog.ProgramName)
                    srec = student_year_records.get(prog.id)
                    frec = staff_year_records.get(prog.id)
                else:
                    ws_students.cell(cur_row_s, 14, "No Program")
                    ws_staff.cell(cur_row_f, 14, "No Program")
                    srec = None
                    frec = None

                # STUDENT: write student_fields starting at col 15
                for i, field in enumerate(student_fields):
                    val = 0
                    if srec:
                        val = getattr(srec, field, None)
                        if val is None:
                            val = getattr(srec, field, 0) or 0
                        val = val or 0
                    ws_students.cell(cur_row_s, 15 + i, val)
                    overall_student_agg[field] += (val or 0)

                # STAFF: write staff_fields starting at col 15
                for j, field in enumerate(staff_fields):
                    val = 0
                    if frec:
                        if hasattr(frec, field):
                            val = getattr(frec, field, 0) or 0
                        else:
                            fallback_map = {
                                "total_staff_washrooms": "total_washrooms",
                                "male_staff_washrooms": "male_washrooms",
                                "female_staff_washrooms": "female_washrooms",
                                "total_staff": "total_staff" if hasattr(frec, "total_staff") else "total_students",
                            }
                            fb = fallback_map.get(field)
                            if fb and hasattr(frec, fb):
                                val = getattr(frec, fb, 0) or 0
                            else:
                                val = getattr(frec, field, 0) or 0
                    ws_staff.cell(cur_row_f, 15 + j, val)
                    overall_staff_agg[field] += (val or 0)

                cur_row_s += 1
                cur_row_f += 1

        row_students = end_row + 1
        row_staff = end_row_s + 1

    # --- Aggregate row per sheet ---
    agg_row = row_students
    ws_students.merge_cells(start_row=agg_row, start_column=1, end_row=agg_row, end_column=12)
    lbl = ws_students.cell(agg_row, 1, f"Aggregate Values - {year}")
    lbl.font = Font(bold=True)
    lbl.alignment = Alignment(horizontal="center", vertical="center")

    ws_students.cell(agg_row, 15, overall_student_agg.get("total_washrooms", 0))
    ws_students.cell(agg_row, 16, overall_student_agg.get("male_washrooms", 0))
    ws_students.cell(agg_row, 17, overall_student_agg.get("female_washrooms", 0))
    ws_students.cell(agg_row, 18, overall_student_agg.get("total_students", 0))
    for col_idx in (15, 16, 17, 18):
        c = ws_students.cell(agg_row, col_idx)
        c.font = Font(bold=True, color="CC6600")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, field in enumerate(student_fields):
        tot = overall_student_agg[field]
        c = ws_students.cell(agg_row, 15 + i, tot)
        c.font = Font(bold=True, color="CC6600")
        c.alignment = Alignment(horizontal="center", vertical="center")

    agg_row_s = row_staff
    ws_staff.merge_cells(start_row=agg_row_s, start_column=1, end_row=agg_row_s, end_column=12)
    lbl2 = ws_staff.cell(agg_row_s, 1, f"Aggregate Values - {year}")
    lbl2.font = Font(bold=True)
    lbl2.alignment = Alignment(horizontal="center", vertical="center")

    ws_staff.cell(agg_row_s, 15, overall_staff_agg.get("total_staff_washrooms", 0))
    ws_staff.cell(agg_row_s, 16, overall_staff_agg.get("male_staff_washrooms", 0))
    ws_staff.cell(agg_row_s, 17, overall_staff_agg.get("female_staff_washrooms", 0))
    ws_staff.cell(agg_row_s, 18, overall_staff_agg.get("total_staff", 0))
    for col_idx in (15, 16, 17, 18):
        c = ws_staff.cell(agg_row_s, col_idx)
        c.font = Font(bold=True, color="CC6600")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for j, field in enumerate(staff_fields):
        tot = overall_staff_agg[field]
        c = ws_staff.cell(agg_row_s, 15 + j, tot)
        c.font = Font(bold=True, color="CC6600")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-width
    students_cols = 14 + len(student_fields)
    for col in range(1, students_cols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws_students[letter]:
            if cell.value is not None:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
        ws_students.column_dimensions[letter].width = min(max_len + 5, 60)

    staff_cols = 14 + len(staff_fields)
    for col in range(1, staff_cols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws_staff[letter]:
            if cell.value is not None:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
        ws_staff.column_dimensions[letter].width = min(max_len + 5, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    date_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Statistics_data_{year}_{date_str}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp




def get_staff_records(request):
    try:
        draw = int(request.GET.get("draw", 1))
        start = int(request.GET.get("start", 0))
        length = int(request.GET.get("length", 10))
    except ValueError:
        return HttpResponseBadRequest("Invalid paging parameters")

    search_value = request.GET.get("search[value]", "").strip()
    order_col_index = request.GET.get("order[0][column]")
    order_dir = request.GET.get("order[0][dir]", "asc")
    year = request.GET.get("year")

    if not year:
        latest = academic_year.objects.order_by("-Academic_Year").first()
        year = latest.Academic_Year if latest else ""

    # 1) Base queryset: colleges that have staff aggregates in this year
    base_qs = College.objects.filter(
        is_deleted=False,
        staff_aggregates__Academic_Year=year,
        staff_aggregates__is_deleted=False,
    ).distinct()

    # 2) Total before search
    records_total = base_qs.count()

    # 3) Apply global search across College + CollegeProgram
    if search_value:
        colleges_qs = base_qs.filter(
            Q(College_Code__icontains=search_value) |
            Q(College_Name__icontains=search_value) |
            Q(address__icontains=search_value) |
            Q(country__icontains=search_value) |
            Q(state__icontains=search_value) |
            Q(District__icontains=search_value) |
            Q(taluka__icontains=search_value) |
            Q(city__icontains=search_value) |
            Q(pincode__icontains=search_value) |
            Q(college_type__icontains=search_value) |
            Q(belongs_to__icontains=search_value) |
            Q(affiliated__icontains=search_value) |
            # from related CollegeProgram
            Q(college_programs__Discipline__icontains=search_value) |
            Q(college_programs__ProgramName__icontains=search_value)
        ).distinct()
    else:
        colleges_qs = base_qs

    # 4) Filtered count after search
    records_filtered = colleges_qs.count()

    # 5) Ordering
    order_map = {
        "1": "College_Code",
        "2": "College_Name",
    }

    if order_col_index == "4":
        # order by total staff (aggregated per college)
        colleges_qs = colleges_qs.annotate(
            agg_total=Sum(
                "staff_aggregates__total_staff",
                filter=Q(staff_aggregates__Academic_Year=year),
            )
        )
        field = "agg_total"
        if order_dir == "desc":
            field = "-" + field
        colleges_qs = colleges_qs.order_by(field, "College_Name")
    else:
        field = order_map.get(order_col_index, "College_Name")
        if order_dir == "desc":
            field = "-" + field
        colleges_qs = colleges_qs.order_by(field)

    # 6) Pagination
    colleges_page = colleges_qs[start: start + length]

    data = []

    for col in colleges_page:
        pc_qs = (
            staff_master_aggregate.objects
            .filter(College=col, Academic_Year=year, is_deleted=False)
            .select_related("Program")
            .order_by("Program__Discipline", "Program__ProgramName")
        )

        discipline_map = {}
        total_staff_for_college = 0

        for pc in pc_qs:
            prog_obj = pc.Program
            discipline = prog_obj.Discipline if prog_obj else "Unspecified"
            program_name = prog_obj.ProgramName if prog_obj else str(pc.Program_id)

            total_staff_for_college += (pc.total_staff or 0)

            entry = {
                "name": program_name,
                "total_staff": pc.total_staff or 0,
                "gender": {
                    "male": pc.total_male or 0,
                    "female": pc.total_female or 0,
                    "others": pc.total_others or 0,
                },
                "washrooms": {
                    "total_washrooms": pc.total_washrooms or 0,
                    "male_washrooms": pc.male_washrooms or 0,
                    "female_washrooms": pc.female_washrooms or 0,
                },
                "category": {
                    "open": {
                        "male": pc.open_male or 0,
                        "female": pc.open_female or 0,
                        "others": pc.open_others or 0,
                    },
                    "obc": {
                        "male": pc.obc_male or 0,
                        "female": pc.obc_female or 0,
                        "others": pc.obc_others or 0,
                    },
                    "sc": {
                        "male": pc.sc_male or 0,
                        "female": pc.sc_female or 0,
                        "others": pc.sc_others or 0,
                    },
                    "st": {
                        "male": pc.st_male or 0,
                        "female": pc.st_female or 0,
                        "others": pc.st_others or 0,
                    },
                    "nt": {
                        "male": pc.nt_male or 0,
                        "female": pc.nt_female or 0,
                        "others": pc.nt_others or 0,
                    },
                    "vjnt": {
                        "male": pc.vjnt_male or 0,
                        "female": pc.vjnt_female or 0,
                        "others": pc.vjnt_others or 0,
                    },
                    "ews": {
                        "male": pc.ews_male or 0,
                        "female": pc.ews_female or 0,
                        "others": pc.ews_others or 0,
                    },
                },
                "religion": {
                    "hindu": {
                        "male": pc.hindu_male or 0,
                        "female": pc.hindu_female or 0,
                        "others": pc.hindu_others or 0,
                    },
                    "muslim": {
                        "male": pc.muslim_male or 0,
                        "female": pc.muslim_female or 0,
                        "others": pc.muslim_others or 0,
                    },
                    "sikh": {
                        "male": pc.sikh_male or 0,
                        "female": pc.sikh_female or 0,
                        "others": pc.sikh_others or 0,
                    },
                    "christian": {
                        "male": pc.christian_male or 0,
                        "female": pc.christian_female or 0,
                        "others": pc.christian_others or 0,
                    },
                    "jain": {
                        "male": pc.jain_male or 0,
                        "female": pc.jain_female or 0,
                        "others": pc.jain_others or 0,
                    },
                    "buddhist": {
                        "male": pc.buddhist_male or 0,
                        "female": pc.buddhist_female or 0,
                        "others": pc.buddhist_others or 0,
                    },
                    "other": {
                        "male": pc.other_religion_male or 0,
                        "female": pc.other_religion_female or 0,
                        "others": pc.other_religion_others or 0,
                    },
                },
                "disability": {
                    "no_disability": {
                        "male": pc.no_disability_male or 0,
                        "female": pc.no_disability_female or 0,
                        "others": pc.no_disability_others or 0,
                    },
                    "lowvision": {
                        "male": pc.low_vision_male or 0,
                        "female": pc.low_vision_female or 0,
                        "others": pc.low_vision_others or 0,
                    },
                    "blindness": {
                        "male": pc.blindness_male or 0,
                        "female": pc.blindness_female or 0,
                        "others": pc.blindness_others or 0,
                    },
                    "hearing": {
                        "male": pc.hearing_male or 0,
                        "female": pc.hearing_female or 0,
                        "others": pc.hearing_others or 0,
                    },
                    "locomotor": {
                        "male": pc.locomotor_male or 0,
                        "female": pc.locomotor_female or 0,
                        "others": pc.locomotor_others or 0,
                    },
                    "autism": {
                        "male": pc.autism_male or 0,
                        "female": pc.autism_female or 0,
                        "others": pc.autism_others or 0,
                    },
                    "other": {
                        "male": pc.other_disability_male or 0,
                        "female": pc.other_disability_female or 0,
                        "others": pc.other_disability_others or 0,
                    },
                },
            }

            discipline_map.setdefault(discipline, []).append(entry)

        grouped_list = []
        for disc in sorted(discipline_map.keys(), key=str.lower):
            grouped_list.append({
                "discipline": disc,
                "programs": sorted(discipline_map[disc], key=lambda x: x["name"].lower()),
            })

        data.append({
            "college_code": col.College_Code,
            "college_name": col.College_Name,
            "academic_year": year,
            "total_staff": total_staff_for_college,
            "programs": grouped_list,
        })

    return JsonResponse({
        "draw": draw,
        "recordsTotal": records_total,
        "recordsFiltered": records_filtered,
        "data": data,
    })

def add_staff_aggregate(request):
    #backend for adding record to staff aggregate master database
    if request.method != "POST":
        return HttpResponseBadRequest("Only POST allowed")

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    college_code = payload.get('college_code')
    academic_year = payload.get('academic_year')
    records = payload.get('records', {})

    if not college_code or not academic_year:
        return JsonResponse({'status': 400, 'message': 'Missing required fields'}, status=400)

    try:
        college = College.objects.get(College_Code=college_code, is_deleted=False)
    except College.DoesNotExist:
        return JsonResponse({'status': 404, 'message': 'College not found'}, status=404)

    if not isinstance(records, dict) or len(records) == 0:
        return JsonResponse({"status": 400, "message": "No records provided"}, status=400)

    saved = []
    errors = []

    with transaction.atomic():
        for program_name, data in records.items():
            program_obj = CollegeProgram.objects.filter(
                College=college,
                ProgramName=program_name,
                is_deleted=False
            ).first()

            if not program_obj:
                errors.append({
                    "program": program_name,
                    "error": f"Program '{program_name}' not found for college '{college_code}'"
                })
                continue

            # ---- parse numeric fields safely ----
            total_staff = _to_int(data.get("total_staff") or data.get("total_staffs"), 0)

            gender = data.get("gender", {}) or {}
            male = _to_int(gender.get("male"), 0)
            female = _to_int(gender.get("female"), 0)
            others = _to_int(gender.get("others") or gender.get("other"), 0)

            washrooms = data.get("washrooms", {}) or {}
            total_washrooms = _to_int(washrooms.get("total") or washrooms.get("total_washrooms"), 0)
            male_washrooms = _to_int(washrooms.get("male"), 0)
            female_washrooms = _to_int(washrooms.get("female"), 0)

            category = data.get("category", {}) or {}
            open_cat = category.get("open", {}) or {}
            open_male = _to_int(open_cat.get("male"), 0)
            open_female = _to_int(open_cat.get("female"), 0)
            open_others = _to_int(open_cat.get("others") or open_cat.get("other"), 0)

            obc = category.get("obc", {}) or {}
            obc_male = _to_int(obc.get("male"), 0)
            obc_female = _to_int(obc.get("female"), 0)
            obc_others = _to_int(obc.get("others") or obc.get("other"), 0)

            sc = category.get("sc", {}) or {}
            sc_male = _to_int(sc.get("male"), 0)
            sc_female = _to_int(sc.get("female"), 0)
            sc_others = _to_int(sc.get("others") or sc.get("other"), 0)

            st = category.get("st", {}) or {}
            st_male = _to_int(st.get("male"), 0)
            st_female = _to_int(st.get("female"), 0)
            st_others = _to_int(st.get("others") or st.get("other"), 0)

            nt = category.get("nt", {}) or {}
            nt_male = _to_int(nt.get("male"), 0)
            nt_female = _to_int(nt.get("female"), 0)
            nt_others = _to_int(nt.get("others") or nt.get("other"), 0)

            vjnt = category.get("vjnt", {}) or {}
            vjnt_male = _to_int(vjnt.get("male"), 0)
            vjnt_female = _to_int(vjnt.get("female"), 0)
            vjnt_others = _to_int(vjnt.get("others") or vjnt.get("other"), 0)

            ews = category.get("ews", {}) or {}
            ews_male = _to_int(ews.get("male"), 0)
            ews_female = _to_int(ews.get("female"), 0)
            ews_others = _to_int(ews.get("others") or ews.get("other"), 0)

            religion = data.get("religion", {}) or {}
            hindu = religion.get("hindu", {}) or {}
            hindu_male = _to_int(hindu.get("male"), 0)
            hindu_female = _to_int(hindu.get("female"), 0)
            hindu_others = _to_int(hindu.get("others") or hindu.get("other"), 0)

            muslim = religion.get("muslim", {}) or {}
            muslim_male = _to_int(muslim.get("male"), 0)
            muslim_female = _to_int(muslim.get("female"), 0)
            muslim_others = _to_int(muslim.get("others") or muslim.get("other"), 0)

            sikh = religion.get("sikh", {}) or {}
            sikh_male = _to_int(sikh.get("male"), 0)
            sikh_female = _to_int(sikh.get("female"), 0)
            sikh_others = _to_int(sikh.get("others") or sikh.get("other"), 0)

            christian = religion.get("christian", {}) or {}
            christian_male = _to_int(christian.get("male"), 0)
            christian_female = _to_int(christian.get("female"), 0)
            christian_others = _to_int(christian.get("others") or christian.get("other"), 0)

            jain = religion.get("jain", {}) or {}
            jain_male = _to_int(jain.get("male"), 0)
            jain_female = _to_int(jain.get("female"), 0)
            jain_others = _to_int(jain.get("others") or jain.get("other"), 0)

            buddhist = religion.get("buddhist", {}) or {}
            buddhist_male = _to_int(buddhist.get("male"), 0)
            buddhist_female = _to_int(buddhist.get("female"), 0)
            buddhist_others = _to_int(buddhist.get("others") or buddhist.get("other"), 0)

            other_religion = religion.get("other_religion", {}) or {}
            other_religion_male = _to_int(other_religion.get("male"), 0)
            other_religion_female = _to_int(other_religion.get("female"), 0)
            other_religion_others = _to_int(other_religion.get("others") or other_religion.get("other"), 0)

            dis = data.get("disability", {}) or {}
            no_disability = dis.get("no_disability", {}) or {}
            no_disability_male = _to_int(no_disability.get("male"), 0)
            no_disability_female = _to_int(no_disability.get("female"), 0)
            no_disability_others = _to_int(no_disability.get("others") or no_disability.get("other"), 0)

            low_vision = dis.get("lowvision", {}) or {}
            low_vision_male = _to_int(low_vision.get("male"), 0)
            low_vision_female = _to_int(low_vision.get("female"), 0)
            low_vision_others = _to_int(low_vision.get("others") or low_vision.get("other"), 0)

            blindness = dis.get("blindness", {}) or {}
            blindness_male = _to_int(blindness.get("male"), 0)
            blindness_female = _to_int(blindness.get("female"), 0)
            blindness_others = _to_int(blindness.get("others") or blindness.get("other"), 0)

            hearing = dis.get("hearing", {}) or {}
            hearing_male = _to_int(hearing.get("male"), 0)
            hearing_female = _to_int(hearing.get("female"), 0)
            hearing_others = _to_int(hearing.get("others") or hearing.get("other"), 0)

            locomotor = dis.get("locomotor", {}) or {}
            locomotor_male = _to_int(locomotor.get("male"), 0)
            locomotor_female = _to_int(locomotor.get("female"), 0)
            locomotor_others = _to_int(locomotor.get("others") or locomotor.get("other"), 0)

            autism = dis.get("autism", {}) or {}
            autism_male = _to_int(autism.get("male"), 0)
            autism_female = _to_int(autism.get("female"), 0)
            autism_others = _to_int(autism.get("others") or autism.get("other"), 0)

            other_disability = dis.get("other_disability", {}) or {}
            other_disability_male = _to_int(other_disability.get("male"), 0)
            other_disability_female = _to_int(other_disability.get("female"), 0)
            other_disability_others = _to_int(other_disability.get("others") or other_disability.get("other"), 0)

            defaults = {
                "total_staff": total_staff,
                "total_male": male,
                "total_female": female,
                "total_others": others,

                "total_washrooms": total_washrooms,
                "male_washrooms": male_washrooms,
                "female_washrooms": female_washrooms,

                "open_male": open_male,
                "open_female": open_female,
                "open_others": open_others,

                "obc_male": obc_male,
                "obc_female": obc_female,
                "obc_others": obc_others,

                "sc_male": sc_male,
                "sc_female": sc_female,
                "sc_others": sc_others,

                "st_male": st_male,
                "st_female": st_female,
                "st_others": st_others,

                "nt_male": nt_male,
                "nt_female": nt_female,
                "nt_others": nt_others,

                "vjnt_male": vjnt_male,
                "vjnt_female": vjnt_female,
                "vjnt_others": vjnt_others,

                "ews_male": ews_male,
                "ews_female": ews_female,
                "ews_others": ews_others,

                "hindu_male": hindu_male,
                "hindu_female": hindu_female,
                "hindu_others": hindu_others,

                "muslim_male": muslim_male,
                "muslim_female": muslim_female,
                "muslim_others": muslim_others,

                "sikh_male": sikh_male,
                "sikh_female": sikh_female,
                "sikh_others": sikh_others,

                "christian_male": christian_male,
                "christian_female": christian_female,
                "christian_others": christian_others,

                "jain_male": jain_male,
                "jain_female": jain_female,
                "jain_others": jain_others,

                "buddhist_male": buddhist_male,
                "buddhist_female": buddhist_female,
                "buddhist_others": buddhist_others,

                "other_religion_male": other_religion_male,
                "other_religion_female": other_religion_female,
                "other_religion_others": other_religion_others,

                "no_disability_male": no_disability_male,
                "no_disability_female": no_disability_female,
                "no_disability_others": no_disability_others,

                "low_vision_male": low_vision_male,
                "low_vision_female": low_vision_female,
                "low_vision_others": low_vision_others,

                "blindness_male": blindness_male,
                "blindness_female": blindness_female,
                "blindness_others": blindness_others,

                "hearing_male": hearing_male,
                "hearing_female": hearing_female,
                "hearing_others": hearing_others,

                "locomotor_male": locomotor_male,
                "locomotor_female": locomotor_female,
                "locomotor_others": locomotor_others,

                "autism_male": autism_male,
                "autism_female": autism_female,
                "autism_others": autism_others,

                "other_disability_male": other_disability_male,
                "other_disability_female": other_disability_female,
                "other_disability_others": other_disability_others,
            }

            try:
                client_ip = get_client_ip(request)

                # Check only ACTIVE record (is_deleted = False)
                existing = staff_master_aggregate.objects.filter(
                    College=college,
                    Program=program_obj,
                    Academic_Year=academic_year,
                    is_deleted=False,
                ).first()

                if existing:
                    errors.append({
                        "program": program_name,
                        "error": "Record already exists for this college, program and academic year"
                    })
                    continue

                # Create a new staff aggregate record
                obj = staff_master_aggregate.objects.create(
                    College=college,
                    Program=program_obj,
                    Academic_Year=academic_year,
                    is_deleted=False,
                    created_by=client_ip,
                    **defaults,
                )

                saved.append({"program": program_name, "id": obj.pk, "created": True})

            except Exception as e:
                errors.append({"program": program_name, "error": f"DB error: {str(e)}"})
                continue

    response_status = 200 if not errors else 207
    resp = {
        "status": response_status,
        "saved": saved,
        "errors": errors,
        "summary": {
            "created": sum(1 for s in saved if s.get("created")),
            "updated": 0,
            "failed": len(errors)
        }
    }
    return JsonResponse(resp)

def update_staff_aggregate(request):
    #backedn for updating to staff aggregate master database
    if request.method != "POST":
        return HttpResponseBadRequest("Only POST allowed")

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    college_code = payload.get('college_code')
    academic_year = payload.get('academic_year')
    records = payload.get('records', {})

    if not college_code or not academic_year:
        return JsonResponse({'status': 400, 'message': 'Missing required fields'}, status=400)

    try:
        college = College.objects.get(College_Code=college_code, is_deleted=False)
    except College.DoesNotExist:
        return JsonResponse({'status': 404, 'message': 'College not found'}, status=404)

    if not isinstance(records, dict) or len(records) == 0:
        return JsonResponse({"status": 400, "message": "No records provided"}, status=400)

    updated = []
    created = []
    errors = []

    with transaction.atomic():
        for program_name, data in records.items():
            program_obj = CollegeProgram.objects.filter(
                College=college,
                ProgramName=program_name,
                is_deleted=False
            ).first()

            if not program_obj:
                errors.append({
                    "program": program_name,
                    "error": f"Program '{program_name}' not found for college '{college_code}'"
                })
                continue


            total_staff = _to_int(data.get("total_staff") or data.get("total_staffs"), 0)

            gender = data.get("gender", {}) or {}
            male = _to_int(gender.get("male"), 0)
            female = _to_int(gender.get("female"), 0)
            others = _to_int(gender.get("others") or gender.get("other"), 0)

            washrooms = data.get("washrooms", {}) or {}
            total_washrooms = _to_int(washrooms.get("total") or washrooms.get("total_washrooms"), 0)
            male_washrooms = _to_int(washrooms.get("male"), 0)
            female_washrooms = _to_int(washrooms.get("female"), 0)

            category = data.get("category", {}) or {}
            open_cat = category.get("open", {}) or {}
            open_male = _to_int(open_cat.get("male"), 0)
            open_female = _to_int(open_cat.get("female"), 0)
            open_others = _to_int(open_cat.get("others") or open_cat.get("other"), 0)

            obc = category.get("obc", {}) or {}
            obc_male = _to_int(obc.get("male"), 0)
            obc_female = _to_int(obc.get("female"), 0)
            obc_others = _to_int(obc.get("others") or obc.get("other"), 0)

            sc = category.get("sc", {}) or {}
            sc_male = _to_int(sc.get("male"), 0)
            sc_female = _to_int(sc.get("female"), 0)
            sc_others = _to_int(sc.get("others") or sc.get("other"), 0)

            st = category.get("st", {}) or {}
            st_male = _to_int(st.get("male"), 0)
            st_female = _to_int(st.get("female"), 0)
            st_others = _to_int(st.get("others") or st.get("other"), 0)

            nt = category.get("nt", {}) or {}
            nt_male = _to_int(nt.get("male"), 0)
            nt_female = _to_int(nt.get("female"), 0)
            nt_others = _to_int(nt.get("others") or nt.get("other"), 0)

            vjnt = category.get("vjnt", {}) or {}
            vjnt_male = _to_int(vjnt.get("male"), 0)
            vjnt_female = _to_int(vjnt.get("female"), 0)
            vjnt_others = _to_int(vjnt.get("others") or vjnt.get("other"), 0)

            ews = category.get("ews", {}) or {}
            ews_male = _to_int(ews.get("male"), 0)
            ews_female = _to_int(ews.get("female"), 0)
            ews_others = _to_int(ews.get("others") or ews.get("other"), 0)

            religion = data.get("religion", {}) or {}
            hindu = religion.get("hindu", {}) or {}
            hindu_male = _to_int(hindu.get("male"), 0)
            hindu_female = _to_int(hindu.get("female"), 0)
            hindu_others = _to_int(hindu.get("others") or hindu.get("other"), 0)

            muslim = religion.get("muslim", {}) or {}
            muslim_male = _to_int(muslim.get("male"), 0)
            muslim_female = _to_int(muslim.get("female"), 0)
            muslim_others = _to_int(muslim.get("others") or muslim.get("other"), 0)

            sikh = religion.get("sikh", {}) or {}
            sikh_male = _to_int(sikh.get("male"), 0)
            sikh_female = _to_int(sikh.get("female"), 0)
            sikh_others = _to_int(sikh.get("others") or sikh.get("other"), 0)

            christian = religion.get("christian", {}) or {}
            christian_male = _to_int(christian.get("male"), 0)
            christian_female = _to_int(christian.get("female"), 0)
            christian_others = _to_int(christian.get("others") or christian.get("other"), 0)

            jain = religion.get("jain", {}) or {}
            jain_male = _to_int(jain.get("male"), 0)
            jain_female = _to_int(jain.get("female"), 0)
            jain_others = _to_int(jain.get("others") or jain.get("other"), 0)

            buddhist = religion.get("buddhist", {}) or {}
            buddhist_male = _to_int(buddhist.get("male"), 0)
            buddhist_female = _to_int(buddhist.get("female"), 0)
            buddhist_others = _to_int(buddhist.get("others") or buddhist.get("other"), 0)

            other_religion = religion.get("other_religion", {}) or {}
            other_religion_male = _to_int(other_religion.get("male"), 0)
            other_religion_female = _to_int(other_religion.get("female"), 0)
            other_religion_others = _to_int(other_religion.get("others") or other_religion.get("other"), 0)

            dis = data.get("disability", {}) or {}
            no_disability = dis.get("no_disability", {}) or {}
            no_disability_male = _to_int(no_disability.get("male"), 0)
            no_disability_female = _to_int(no_disability.get("female"), 0)
            no_disability_others = _to_int(no_disability.get("others") or no_disability.get("other"), 0)

            low_vision = dis.get("lowvision", {}) or {}
            low_vision_male = _to_int(low_vision.get("male"), 0)
            low_vision_female = _to_int(low_vision.get("female"), 0)
            low_vision_others = _to_int(low_vision.get("others") or low_vision.get("other"), 0)

            blindness = dis.get("blindness", {}) or {}
            blindness_male = _to_int(blindness.get("male"), 0)
            blindness_female = _to_int(blindness.get("female"), 0)
            blindness_others = _to_int(blindness.get("others") or blindness.get("other"), 0)

            hearing = dis.get("hearing", {}) or {}
            hearing_male = _to_int(hearing.get("male"), 0)
            hearing_female = _to_int(hearing.get("female"), 0)
            hearing_others = _to_int(hearing.get("others") or hearing.get("other"), 0)

            locomotor = dis.get("locomotor", {}) or {}
            locomotor_male = _to_int(locomotor.get("male"), 0)
            locomotor_female = _to_int(locomotor.get("female"), 0)
            locomotor_others = _to_int(locomotor.get("others") or locomotor.get("other"), 0)

            autism = dis.get("autism", {}) or {}
            autism_male = _to_int(autism.get("male"), 0)
            autism_female = _to_int(autism.get("female"), 0)
            autism_others = _to_int(autism.get("others") or autism.get("other"), 0)

            other_disability = dis.get("other_disability", {}) or {}
            other_disability_male = _to_int(other_disability.get("male"), 0)
            other_disability_female = _to_int(other_disability.get("female"), 0)
            other_disability_others = _to_int(other_disability.get("others") or other_disability.get("other"), 0)

            try:
                client_ip = get_client_ip(request)

                existing = staff_master_aggregate.objects.filter(
                    College=college,
                    Program=program_obj,
                    Academic_Year=academic_year,
                    is_deleted=False,
                ).first()

                if existing:
                    existing.total_staff = total_staff
                    existing.total_male = male
                    existing.total_female = female
                    existing.total_others = others

                    existing.total_washrooms = total_washrooms
                    existing.male_washrooms = male_washrooms
                    existing.female_washrooms = female_washrooms

                    existing.open_male = open_male
                    existing.open_female = open_female
                    existing.open_others = open_others

                    existing.obc_male = obc_male
                    existing.obc_female = obc_female
                    existing.obc_others = obc_others

                    existing.sc_male = sc_male
                    existing.sc_female = sc_female
                    existing.sc_others = sc_others

                    existing.st_male = st_male
                    existing.st_female = st_female
                    existing.st_others = st_others

                    existing.nt_male = nt_male
                    existing.nt_female = nt_female
                    existing.nt_others = nt_others

                    existing.vjnt_male = vjnt_male
                    existing.vjnt_female = vjnt_female
                    existing.vjnt_others = vjnt_others

                    existing.ews_male = ews_male
                    existing.ews_female = ews_female
                    existing.ews_others = ews_others

                    existing.hindu_male = hindu_male
                    existing.hindu_female = hindu_female
                    existing.hindu_others = hindu_others

                    existing.muslim_male = muslim_male
                    existing.muslim_female = muslim_female
                    existing.muslim_others = muslim_others

                    existing.sikh_male = sikh_male
                    existing.sikh_female = sikh_female
                    existing.sikh_others = sikh_others

                    existing.christian_male = christian_male
                    existing.christian_female = christian_female
                    existing.christian_others = christian_others

                    existing.jain_male = jain_male
                    existing.jain_female = jain_female
                    existing.jain_others = jain_others

                    existing.buddhist_male = buddhist_male
                    existing.buddhist_female = buddhist_female
                    existing.buddhist_others = buddhist_others

                    existing.other_religion_male = other_religion_male
                    existing.other_religion_female = other_religion_female
                    existing.other_religion_others = other_religion_others

                    existing.no_disability_male = no_disability_male
                    existing.no_disability_female = no_disability_female
                    existing.no_disability_others = no_disability_others

                    existing.low_vision_male = low_vision_male
                    existing.low_vision_female = low_vision_female
                    existing.low_vision_others = low_vision_others

                    existing.blindness_male = blindness_male
                    existing.blindness_female = blindness_female
                    existing.blindness_others = blindness_others

                    existing.hearing_male = hearing_male
                    existing.hearing_female = hearing_female
                    existing.hearing_others = hearing_others

                    existing.locomotor_male = locomotor_male
                    existing.locomotor_female = locomotor_female
                    existing.locomotor_others = locomotor_others

                    existing.autism_male = autism_male
                    existing.autism_female = autism_female
                    existing.autism_others = autism_others

                    existing.other_disability_male = other_disability_male
                    existing.other_disability_female = other_disability_female
                    existing.other_disability_others = other_disability_others

                    existing.updated_by = client_ip
                    existing.save()

                    updated.append({"program": program_name, "id": existing.pk, "updated": True})
                else:
                    # Create a new staff aggregate record (uses total_staff)
                    obj = staff_master_aggregate.objects.create(
                        College=college,
                        Program=program_obj,
                        Academic_Year=academic_year,
                        is_deleted=False,
                        created_by=client_ip,

                        total_staff=total_staff,
                        total_male=male,
                        total_female=female,
                        total_others=others,

                        total_washrooms=total_washrooms,
                        male_washrooms=male_washrooms,
                        female_washrooms=female_washrooms,

                        open_male=open_male,
                        open_female=open_female,
                        open_others=open_others,

                        obc_male=obc_male,
                        obc_female=obc_female,
                        obc_others=obc_others,

                        sc_male=sc_male,
                        sc_female=sc_female,
                        sc_others=sc_others,

                        st_male=st_male,
                        st_female=st_female,
                        st_others=st_others,

                        nt_male=nt_male,
                        nt_female=nt_female,
                        nt_others=nt_others,

                        vjnt_male=vjnt_male,
                        vjnt_female=vjnt_female,
                        vjnt_others=vjnt_others,

                        ews_male=ews_male,
                        ews_female=ews_female,
                        ews_others=ews_others,

                        hindu_male=hindu_male,
                        hindu_female=hindu_female,
                        hindu_others=hindu_others,

                        muslim_male=muslim_male,
                        muslim_female=muslim_female,
                        muslim_others=muslim_others,

                        sikh_male=sikh_male,
                        sikh_female=sikh_female,
                        sikh_others=sikh_others,

                        christian_male=christian_male,
                        christian_female=christian_female,
                        christian_others=christian_others,

                        jain_male=jain_male,
                        jain_female=jain_female,
                        jain_others=jain_others,

                        buddhist_male=buddhist_male,
                        buddhist_female=buddhist_female,
                        buddhist_others=buddhist_others,

                        other_religion_male=other_religion_male,
                        other_religion_female=other_religion_female,
                        other_religion_others=other_religion_others,

                        no_disability_male=no_disability_male,
                        no_disability_female=no_disability_female,
                        no_disability_others=no_disability_others,

                        low_vision_male=low_vision_male,
                        low_vision_female=low_vision_female,
                        low_vision_others=low_vision_others,

                        blindness_male=blindness_male,
                        blindness_female=blindness_female,
                        blindness_others=blindness_others,

                        hearing_male=hearing_male,
                        hearing_female=hearing_female,
                        hearing_others=hearing_others,

                        locomotor_male=locomotor_male,
                        locomotor_female=locomotor_female,
                        locomotor_others=locomotor_others,

                        autism_male=autism_male,
                        autism_female=autism_female,
                        autism_others=autism_others,

                        other_disability_male=other_disability_male,
                        other_disability_female=other_disability_female,
                        other_disability_others=other_disability_others,
                    )
                    created.append({"program": program_name, "id": obj.pk, "created": True})

            except Exception as e:
                errors.append({"program": program_name, "error": f"DB error: {str(e)}"})
                continue

    response_status = 200 if not errors else 207
    resp = {
        "status": response_status,
        "created": created,
        "updated": updated,
        "errors": errors,
        "summary": {
            "created": len(created),
            "updated": len(updated),
            "failed": len(errors)
        }
    }
    return JsonResponse(resp, status=response_status)



def delete_staff_record(request):
    #soft delete staff aggregate same as students
    if request.method == 'POST':
        college_code = request.POST.get('college_code')
        academic_year = request.POST.get('academic_year')

        try:
            college = College.objects.get(College_Code=college_code, is_deleted=False)
        except College.DoesNotExist:
            return JsonResponse({'status': 404, 'message': 'College not found'})

        staff_master_aggregate.objects.filter(College=college, Academic_Year=academic_year, is_deleted=False).update(is_deleted=True)

        response_data = {
            'message': 'Staff records deleted successfully',
            'status': 204
        }
        return JsonResponse(response_data)


@ajax_login_required
def export_staff_excel(request):
    if request.method != "POST":
        return HttpResponseBadRequest("Only POST allowed")

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return HttpResponseBadRequest("Invalid JSON payload")

    year = payload.get("year")
    if not year:
        return HttpResponseBadRequest("Missing academic year")

    global_search = (payload.get("search") or "").strip()
    order_instructions = payload.get("order", []) or []

    # -----------------------
    # Base queryset (staff aggregates)
    # -----------------------
    qs = College.objects.filter(
        is_deleted=False,
        staff_aggregates__Academic_Year=year,
        staff_aggregates__is_deleted=False
    ).prefetch_related("college_programs", "staff_aggregates").distinct()

    # -----------------------
    # Global Search
    # -----------------------
    if global_search:
        qs = qs.filter(
            Q(College_Name__icontains=global_search)
            | Q(College_Code__icontains=global_search)
            | Q(address__icontains=global_search)
            | Q(country__icontains=global_search)
            | Q(state__icontains=global_search)
            | Q(District__icontains=global_search)
            | Q(taluka__icontains=global_search)
            | Q(city__icontains=global_search)
            | Q(pincode__icontains=global_search)
            | Q(college_programs__ProgramName__icontains=global_search)
            | Q(college_programs__Discipline__icontains=global_search)
        ).distinct()

    # -----------------------
    # Ordering support
    # -----------------------
    COLUMN_INDEX_TO_FIELD = {
        1: "College_Code",
        2: "College_Name",
        3: "staff_aggregates__Academic_Year",
        4: "staff_aggregates__total_staff",
    }

    order_by = []
    try:
        for pair in order_instructions:
            if isinstance(pair, (list, tuple)) and len(pair) >= 2:
                cidx = int(pair[0])
                direction = (pair[1] or "asc").lower()
            elif isinstance(pair, dict):
                cidx = int(pair.get("column") or pair.get("col") or pair.get("0", 0))
                direction = (pair.get("dir") or "asc").lower()
            else:
                continue

            field = COLUMN_INDEX_TO_FIELD.get(cidx)
            if field:
                order_by.append(("-" if direction == "desc" else "") + field)
    except Exception:
        pass

    if order_by:
        qs = qs.order_by(*order_by)

    # -----------------------
    # Build XLSX
    # -----------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Records"

    # HEADERS
    headers = [
        "College Code", "College Name", "Address", "Pincode", "Country",
        "State", "District", "Taluka", "City",
        "College Type", "Belongs To", "Affiliated To",

        "Discipline", "Program",

        "Total Staff Washrooms",
        "Male Staff Washrooms",
        "Female Staff Washrooms",

        "Total Staff",
        "Total Male", "Total Female", "Total Others",

        # caste
        "OPEN Male", "OPEN Female", "OPEN Others",
        "OBC Male", "OBC Female", "OBC Others",
        "SC Male", "SC Female", "SC Others",
        "ST Male", "ST Female", "ST Others",
        "NT Male", "NT Female", "NT Others",
        "VJNT Male", "VJNT Female", "VJNT Others",
        "EWS Male", "EWS Female", "EWS Others",

        # religion
        "Hindu Male", "Hindu Female", "Hindu Others",
        "Muslim Male", "Muslim Female", "Muslim Others",
        "Sikh Male", "Sikh Female", "Sikh Others",
        "Christian Male", "Christian Female", "Christian Others",
        "Jain Male", "Jain Female", "Jain Others",
        "Buddhist Male", "Buddhist Female", "Buddhist Others",
        "Other Religion Male", "Other Religion Female", "Other Religion Others",

        # disability
        "No Disability Male", "No Disability Female", "No Disability Others",
        "Low Vision Male", "Low Vision Female", "Low Vision Others",
        "Blindness Male", "Blindness Female", "Blindness Others",
        "Hearing Impaired Male", "Hearing Impaired Female", "Hearing Impaired Others",
        "Locomotor Disability Male", "Locomotor Disability Female", "Locomotor Disability Others",
        "Autism Male", "Autism Female", "Autism Others",
        "Other Disability Male", "Other Disability Female", "Other Disability Others",
    ]
    ws.append(headers)

    # style header
    header_fill = PatternFill(start_color="006699", fill_type="solid")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # STAFF FIELDS
    staff_fields = [
        "total_staff_washrooms", "male_staff_washrooms", "female_staff_washrooms",
        "total_staff",
        "total_male", "total_female", "total_others",

        "open_male", "open_female", "open_others",
        "obc_male", "obc_female", "obc_others",
        "sc_male", "sc_female", "sc_others",
        "st_male", "st_female", "st_others",
        "nt_male", "nt_female", "nt_others",
        "vjnt_male", "vjnt_female", "vjnt_others",
        "ews_male", "ews_female", "ews_others",

        "hindu_male", "hindu_female", "hindu_others",
        "muslim_male", "muslim_female", "muslim_others",
        "sikh_male", "sikh_female", "sikh_others",
        "christian_male", "christian_female", "christian_others",
        "jain_male", "jain_female", "jain_others",
        "buddhist_male", "buddhist_female", "buddhist_others",
        "other_religion_male", "other_religion_female", "other_religion_others",

        "no_disability_male", "no_disability_female", "no_disability_others",
        "low_vision_male", "low_vision_female", "low_vision_others",
        "blindness_male", "blindness_female", "blindness_others",
        "hearing_male", "hearing_female", "hearing_others",
        "locomotor_male", "locomotor_female", "locomotor_others",
        "autism_male", "autism_female", "autism_others",
        "other_disability_male", "other_disability_female", "other_disability_others",
    ]

    overall_agg = {f: 0 for f in staff_fields}

    row_num = 2

    # -----------------------
    # MAIN LOOP
    # -----------------------
    for college in qs:

        # year-based staff records
        year_records = {
            r.Program_id: r
            for r in college.staff_aggregates.filter(Academic_Year=year, is_deleted=False)
        }

        # filter programs based on global_search
        master_programs = college.college_programs.filter(is_deleted=False)

        if global_search:
            allowed_programs = list(master_programs.filter(
                Q(Discipline__icontains=global_search) |
                Q(ProgramName__icontains=global_search)
            ))

            if not allowed_programs:
                continue
        else:
            allowed_programs = list(master_programs)

        if not allowed_programs:
            allowed_programs = [None]

        # group by discipline
        discipline_map = {}
        for cp in allowed_programs:
            if cp is None:
                discipline_map.setdefault("No Discipline", []).append(None)
            else:
                discipline_map.setdefault(cp.Discipline or "Unspecified", []).append(cp)

        discipline_list = list(discipline_map.items())

        total_program_rows = sum(len(v) if v else 1 for _, v in discipline_list)
        if total_program_rows <= 0:
            total_program_rows = 1

        start_row = row_num
        end_row = row_num + total_program_rows - 1

        # merged college fields
        college_fields = [
            college.College_Code,
            college.College_Name,
            college.address or "",
            college.pincode or "",
            college.country or "",
            college.state or "",
            college.District or "",
            college.taluka or "",
            college.city or "",
            college.college_type or "",
            college.belongs_to or "",
            college.affiliated or "",
        ]

        for ci, val in enumerate(college_fields, start=1):
            ws.merge_cells(
                start_row=start_row,
                start_column=ci,
                end_row=end_row,
                end_column=ci
            )
            cell = ws.cell(row=start_row, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        # discipline + program listing
        current_row = row_num
        for discipline, progs in discipline_list:

            if not progs:
                progs = [None]

            ws.merge_cells(
                start_row=current_row,
                start_column=13,
                end_row=current_row + len(progs) - 1,
                end_column=13
            )
            ws.cell(row=current_row, column=13, value=discipline).alignment = Alignment(
                vertical="center", horizontal="center", wrap_text=True
            )

            for cp in progs:
                if cp:
                    ws.cell(row=current_row, column=14, value=cp.ProgramName)
                    rec = year_records.get(cp.id)
                else:
                    ws.cell(row=current_row, column=14, value="No Program")
                    rec = None

                # write staff fields
                for i, field in enumerate(staff_fields):
                    if rec and hasattr(rec, field):
                        v = getattr(rec, field) or 0
                    else:
                        # fallback mapping for legacy names
                        fallback = {
                            "total_staff_washrooms": "total_washrooms",
                            "male_staff_washrooms": "male_washrooms",
                            "female_staff_washrooms": "female_washrooms",
                        }
                        v = getattr(rec, fallback.get(field, field), 0) if rec else 0

                    ws.cell(row=current_row, column=15 + i, value=v)
                    overall_agg[field] += v

                current_row += 1

        row_num = end_row + 1

    # -----------------------
    # FINAL TOTAL ROW
    # -----------------------
    agg_row = row_num

    ws.merge_cells(start_row=agg_row, start_column=1, end_row=agg_row, end_column=12)
    label = ws.cell(agg_row, 1, f"Aggregate Values - {year}")
    label.font = Font(bold=True)
    label.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(agg_row, 15, overall_agg["total_staff_washrooms"])
    ws.cell(agg_row, 16, overall_agg["male_staff_washrooms"])
    ws.cell(agg_row, 17, overall_agg["female_staff_washrooms"])
    ws.cell(agg_row, 18, overall_agg["total_staff"])

    for col_idx in (15, 16, 17, 18):
        c = ws.cell(agg_row, col_idx)
        c.font = Font(bold=True, color="CC6600")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, field in enumerate(staff_fields):
        val = overall_agg[field]
        c = ws.cell(agg_row, 15 + i, value=val)
        c.font = Font(bold=True, color="CC6600")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # auto-size columns
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 5, 60)

    # respond
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Staff_Report_{year}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return resp


User = get_user_model()


def unassigned_users_json(request):

    if request.method == "GET":

         college_id = request.GET.get("college_id")

    current_user = None
    current_user_id = None

    # 1) In edit mode, find current user assigned to this college (if any)
    if college_id:
        profile = (
            UserCollege.objects
            .filter(college_id=college_id)
            .select_related("user")
            .first()
        )
        if profile and profile.user:
            current_user = profile.user
            current_user_id = current_user.id

    # 2) Base: all active, non-superuser users
    base_users = User.objects.filter(is_active=True, is_superuser=False)

    # 3) Users already assigned to some college
    assigned_user_ids = list(
        UserCollege.objects
        .filter(college__isnull=False)
        .values_list("user_id", flat=True)
    )

    # 4) Unassigned = base - assigned
    unassigned_users_qs = base_users.exclude(id__in=assigned_user_ids)

    result = []

    # 5) If we're editing and the college already has a user, include that user first
    if current_user:
        result.append({
            "id": current_user.id,
            "username": current_user.username,
            "is_current": True,
        })

    # 6) Add all unassigned users
    for u in unassigned_users_qs.order_by("username"):
        result.append({
            "id": u.id,
            "username": u.username,
            "is_current": False,
        })

    return JsonResponse({
        "users": result,
        "current_user_id": current_user_id,
    })

        


    return JsonResponse({"users": result})


def change_password(request):
    user = request.user

    old = request.POST.get("old_password", "").strip()
    new = request.POST.get("new_password", "").strip()

    if not old or not new:
        return JsonResponse({"error": "Both old and new passwords are required."}, status=400)

    # verify current password
    if not user.check_password(old):
        return JsonResponse({"error": "Old password is incorrect."}, status=403)

    # set new password and save
    user.set_password(new)
    user.save()

    # Note: by default Django does not force logout on password change.
    # If you WANT to keep the user logged-in you can call:
    #   update_session_auth_hash(request, user)
    # But per your instruction, leaving default behavior is fine.

    return JsonResponse({"success": True})
